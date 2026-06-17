#!/usr/bin/env python3
"""金融 RVEC 综合打标 —— 自包含单文件脚本（纯标准库 + openpyxl 版）

仅依赖: Python 标准库 + openpyxl（环境已有，直接使用不经过 pandas）
LLM 调用使用 urllib.request 直连 OpenAI 兼容 API，无需 openai SDK。
不依赖 pandas，避免 pandas 对 openpyxl 版本的要求。

用法：
  # 数据探查（不调用 LLM，仅分析数据结构）
  python3 fin_rvec_tag.py --input data.xlsx --inspect

  # 预览打标（打标前3条，输出对比JSON）
  python3 fin_rvec_tag.py --input data.xlsx --preview 3

  # 正式打标
  python3 fin_rvec_tag.py --input data.xlsx --output output/

  # 指定端点
  python3 fin_rvec_tag.py --input data.xlsx --endpoint iquest
  python3 fin_rvec_tag.py --input data.xlsx --endpoint zerail

  # 自定义 LLM
  python3 fin_rvec_tag.py --input data.xlsx --base-url http://xxx/v1 --api-key sk-xxx --model gpt-4o
"""

from __future__ import annotations

import argparse
import concurrent.futures
import csv
import json
import logging
import os
import re
import socket
import ssl
import sys
import time
import random
import threading
import urllib.request
import urllib.error
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# ════════════════════════════════════════════════════════════
# 1. LLM 客户端（纯 urllib 实现，无需 openai SDK）
# ════════════════════════════════════════════════════════════

DEFAULT_ENDPOINTS = {
    "iquest": {
        "base_url": "http://iqeust-litellm.danbo-agidata-inner.com/v1",
        "api_key": "sk-kOp807j6jMPRirtnj9bAJg",
        "model": "gpt-5.4",
    },
    "zerail": {
        "base_url": "https://gateway.zerail.com/v1",
        "api_key": "sk-eModH1YZpV9YVdvc1WLA5mEvwFYbEkrKbSJq0TUbxWKe2y1K",
        "model": "gpt-5.2",
    },
}

# 宽松 SSL 上下文（部分内网端点证书不标准）
_SSL_CTX = ssl.create_default_context()
_SSL_CTX.check_hostname = False
_SSL_CTX.verify_mode = ssl.CERT_NONE


class LLMClient:
    """纯 urllib 实现的 OpenAI 兼容客户端"""

    def __init__(self, api_key=None, base_url=None, model="gpt-4o", temperature=0.0,
                 endpoint_name=None):
        self.temperature = temperature
        if endpoint_name and endpoint_name in DEFAULT_ENDPOINTS:
            cfg = DEFAULT_ENDPOINTS[endpoint_name]
            api_key, base_url, model = cfg["api_key"], cfg["base_url"], cfg.get("model", model)
        self.model = model
        self.api_key = api_key or os.getenv("OPENAI_API_KEY", "")
        self.base_url = (base_url or os.getenv("OPENAI_BASE_URL", "")).rstrip("/")

    @classmethod
    def auto_detect(cls, temperature=0.0) -> "LLMClient":
        """自动探测可用端点：DNS预检 → chat ping"""
        for name, cfg in DEFAULT_ENDPOINTS.items():
            try:
                host = urlparse(cfg["base_url"]).hostname
                old_timeout = socket.getdefaulttimeout()
                socket.setdefaulttimeout(5)
                try:
                    socket.getaddrinfo(host, 80)
                finally:
                    socket.setdefaulttimeout(old_timeout)
                # chat ping
                client = cls(api_key=cfg["api_key"], base_url=cfg["base_url"],
                             model=cfg["model"], temperature=temperature)
                client._raw_chat([{"role": "user", "content": "ping"}], max_tokens=3, timeout=10)
                logger.info(f"✅ 端点 [{name}] 可用 | 模型: {cfg['model']} @ {host}")
                return client
            except Exception as e:
                logger.warning(f"端点 [{name}] 不可用: {str(e)[:120]}")
        raise RuntimeError("所有 LLM 端点均不可达")

    def _raw_chat(self, messages: list, max_tokens=4096, timeout=300) -> str:
        """直接用 urllib 调用 OpenAI 兼容 /chat/completions"""
        url = f"{self.base_url}/chat/completions"
        payload = json.dumps({
            "model": self.model,
            "temperature": self.temperature,
            "messages": messages,
            "max_tokens": max_tokens,
        }).encode("utf-8")
        req = urllib.request.Request(
            url, data=payload,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {self.api_key}",
            },
        )
        handler = urllib.request.HTTPSHandler(context=_SSL_CTX) if url.startswith("https") else urllib.request.HTTPHandler()
        opener = urllib.request.build_opener(handler)
        resp = opener.open(req, timeout=timeout)
        body = json.loads(resp.read().decode("utf-8"))
        return body["choices"][0]["message"]["content"]

    def chat(self, system_prompt: str, user_prompt: str, retries=5) -> str:
        """带重试的 chat 调用（对 429 限流做长退避）"""
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]
        last_err = None
        for attempt in range(retries):
            try:
                return self._raw_chat(messages)
            except urllib.error.HTTPError as e:
                last_err = e
                if e.code == 429:
                    # 限流：指数退避 5s, 15s, 30s, 60s, 60s
                    wait = min(5 * (3 ** attempt), 60) + random.uniform(1, 3)
                else:
                    wait = min(2 ** attempt, 30) + random.uniform(0, 1)
                logger.warning(f"LLM 调用失败 (attempt {attempt+1}/{retries}): {str(e)[:100]}，{wait:.1f}s 后重试")
                time.sleep(wait)
            except Exception as e:
                last_err = e
                wait = min(2 ** attempt, 30) + random.uniform(0, 1)
                logger.warning(f"LLM 调用失败 (attempt {attempt+1}/{retries}): {str(e)[:100]}，{wait:.1f}s 后重试")
                time.sleep(wait)
        raise RuntimeError(f"LLM 调用失败（{retries}次重试后）: {last_err}")

    def chat_json_with_raw(self, system_prompt: str, user_prompt: str):
        """调 chat 拿到 raw 字符串 + 解析后 JSON。供 explain 模式使用。"""
        raw = self.chat(system_prompt, user_prompt)
        cleaned = raw
        if "```json" in cleaned:
            cleaned = cleaned.split("```json")[1].split("```")[0]
        elif "```" in cleaned:
            cleaned = cleaned.split("```")[1].split("```")[0]
        parsed = json.loads(cleaned.strip())
        return raw, parsed

    def chat_json(self, system_prompt: str, user_prompt: str) -> dict:
        _raw, parsed = self.chat_json_with_raw(system_prompt, user_prompt)
        return parsed


# ════════════════════════════════════════════════════════════
# 2. IO 工具（纯 openpyxl + csv 标准库，不依赖 pandas）
# ════════════════════════════════════════════════════════════

def read_data(file_path: str) -> List[Dict[str, Any]]:
    p = Path(file_path)
    ext = p.suffix.lower()
    if ext in (".xlsx", ".xls"):
        return _read_xlsx(file_path)
    elif ext == ".csv":
        return _read_csv(file_path)
    elif ext in (".jsonl", ".json"):
        records = []
        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    records.append(json.loads(line))
        return records
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def _read_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """直接用 openpyxl 读取 xlsx，不经过 pandas（避免版本冲突）"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data_rows = rows[1:]
    # 检测第一数据行是否为描述行（全部长文本）
    if len(data_rows) > 1:
        first = data_rows[0]
        non_none = [v for v in first if v is not None]
        if non_none and all(isinstance(v, str) and len(v) > 30 for v in non_none):
            data_rows = data_rows[1:]
    records = []
    for row in data_rows:
        record = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            record[h] = str(val) if val is not None else ""
        records.append(record)
    return records


def _read_csv(file_path: str) -> List[Dict[str, Any]]:
    """用标准库 csv 读取（提高 field_size_limit 以容纳超长 model_response）"""
    # 解除 csv 默认 131072 字段大小限制（report 池含超长字段）
    try:
        csv.field_size_limit(sys.maxsize)
    except OverflowError:
        # Windows 上 sys.maxsize 可能超过 C long，降级到 INT32 上限
        csv.field_size_limit(2**31 - 1)
    records = []
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append({k: (v if v is not None else "") for k, v in row.items()})
    return records


def _flatten_value(v):
    """将 dict/list 转为 JSON 字符串"""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return v


def write_data(records: List[Dict], file_path: str, fmt: str):
    if not records:
        return
    p = Path(file_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    if fmt == "xlsx":
        _write_xlsx(records, str(p.with_suffix(".xlsx")))
    elif fmt == "csv":
        _write_csv(records, str(p))
    else:  # jsonl
        with open(str(p), "w", encoding="utf-8") as f:
            for r in records:
                f.write(json.dumps(r, ensure_ascii=False) + "\n")
        logger.info(f"📄 写入: {p} ({len(records)} 行)")


def _collect_all_headers(records: List[Dict]) -> List[str]:
    """收集所有记录的列名并按出现顺序合并去重，避免第一条缺字段时丢列。"""
    seen = set()
    headers = []
    for r in records:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                headers.append(k)
    return headers


def _write_xlsx(records: List[Dict], file_path: str):
    """直接用 openpyxl 写入 xlsx"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    if not records:
        wb.save(file_path)
        return
    headers = _collect_all_headers(records)
    ws.append(headers)
    for r in records:
        ws.append([_flatten_value(r.get(h, "")) for h in headers])
    wb.save(file_path)
    logger.info(f"📄 写入: {file_path} ({len(records)} 行, {len(headers)} 列)")


def _write_csv(records: List[Dict], file_path: str):
    """用标准库 csv 写入"""
    if not records:
        return
    headers = _collect_all_headers(records)
    with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for r in records:
            writer.writerow({k: _flatten_value(r.get(k, "")) for k in headers})
    logger.info(f"📄 写入: {file_path} ({len(records)} 行, {len(headers)} 列)")


# ════════════════════════════════════════════════════════════
# 3. RVEC 配置 & Prompt
# ════════════════════════════════════════════════════════════

def _read_yaml_or_json(p: Path) -> Optional[Dict]:
    """读取单个 yaml/json 文件，文件不存在或解析失败返回 None。
    yaml 不可用时静默回退（不报错），让调用方决定要不要用 fallback。
    """
    if not p.exists():
        return None
    try:
        text = p.read_text(encoding="utf-8")
    except OSError:
        return None
    # 优先按 JSON 解析（无需 pyyaml）
    if p.suffix.lower() == ".json":
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return None
    # YAML 路径
    try:
        import yaml
        return yaml.safe_load(text)
    except ImportError:
        logger.warning(f"⚠️  pyyaml 未安装，跳过 {p.name}（如需启用请 pip install pyyaml）")
        return None
    except Exception as e:
        logger.warning(f"⚠️  解析 {p.name} 失败: {e}")
        return None


def _load_config(config_path: Optional[str]) -> Dict:
    """加载配置文件（多源合并 + 自动回退）。

    加载优先级：
      1. 用户显式 --config 指定 → 直接用该文件（不 merge rvec/）
      2. 默认情况：先读 base config（fin_rvec_config.json/yaml），
         再用 config/rvec/{prompts,families,rules}.yaml 覆盖（如果存在）

    rvec/ 下的 3 个文件均为可选，缺失则回退到 base config 中的对应字段。
    这样保证：
      - 服务器上没有 rvec/ 目录时跑批不受影响
      - 用户改 rvec/prompts.yaml 等不需要动 fin_rvec_config.yaml
    """
    base_dir = Path(__file__).resolve().parent.parent / "config"

    # ── 1. 加载 base config ──
    if config_path:
        p = Path(config_path)
        if not p.exists():
            raise FileNotFoundError(f"配置文件不存在: {p}")
        base_config = _read_yaml_or_json(p)
        if base_config is None:
            raise ValueError(f"配置文件解析失败: {p}")
        # 用户显式指定 config，不再 merge rvec/，避免意外覆盖
        return base_config

    # 默认 base：优先 json（无需 pyyaml），其次 yaml
    p_json = base_dir / "fin_rvec_config.json"
    p_yaml = base_dir / "fin_rvec_config.yaml"
    p = p_json if p_json.exists() else p_yaml
    if not p.exists():
        raise FileNotFoundError(f"配置文件不存在: {p}")
    base_config = _read_yaml_or_json(p)
    if base_config is None:
        raise ValueError(
            f"配置文件解析失败: {p}\n"
            f"如果是 YAML 但 pyyaml 未安装，请使用 fin_rvec_config.json 或 pip install pyyaml"
        )

    # ── 2. 用 rvec/*.yaml 覆盖（可选）──
    rvec_dir = base_dir / "rvec"
    if rvec_dir.is_dir():
        # families.yaml → task_family_rules
        families_data = _read_yaml_or_json(rvec_dir / "families.yaml")
        if families_data and "task_family_rules" in families_data:
            base_config["task_family_rules"] = families_data["task_family_rules"]
            logger.debug("✅ 已加载 rvec/families.yaml")

        # rules.yaml → fin_scene_schema / task_type_schema / rvec_schema / severity_schema / scoring_rules / ft_cross_validation
        rules_data = _read_yaml_or_json(rvec_dir / "rules.yaml")
        if rules_data:
            for key in ("fin_scene_schema", "task_type_schema", "rvec_schema",
                        "severity_schema", "scoring_rules", "ft_cross_validation"):
                if key in rules_data:
                    base_config[key] = rules_data[key]
            logger.debug("✅ 已加载 rvec/rules.yaml")

        # prompts.yaml → 注入到 _prompts 子字典（供 _resolve_prompts 取用）
        prompts_data = _read_yaml_or_json(rvec_dir / "prompts.yaml")
        if prompts_data:
            base_config["_prompts"] = prompts_data
            logger.debug("✅ 已加载 rvec/prompts.yaml")

    return base_config


def _resolve_prompts(config: Dict) -> Dict[str, Any]:
    """解析 prompt 来源：优先 config["_prompts"]（YAML），否则回退到代码内置常量。

    返回:
      {
        "system_prompt":     <模板字符串>,
        "user_prompt_header": <模板字符串>,
        "family_templates":   {family: <完整 user prompt 模板>}
      }
    其中 family_templates 中的每个模板已经是 user_prompt_header + family 块的拼接结果。
    """
    yaml_prompts = config.get("_prompts") or {}
    sys_tpl = yaml_prompts.get("system_prompt") or SYSTEM_PROMPT_TEMPLATE
    header_tpl = yaml_prompts.get("user_prompt_header") or USER_PROMPT_HEADER

    yaml_family = yaml_prompts.get("family_templates") or {}
    fallback_family = {
        "QA_CHOICE": USER_PROMPT_QA_CHOICE,
        "SENTIMENT": USER_PROMPT_SENTIMENT,
        "REPORT_EVAL": USER_PROMPT_REPORT_EVAL,
        "LONG_GEN": USER_PROMPT_LONG_GEN,
        "GENERIC": USER_PROMPT_GENERIC,
    }

    merged_family = {}
    for fam in ("QA_CHOICE", "SENTIMENT", "REPORT_EVAL", "LONG_GEN", "GENERIC"):
        body = yaml_family.get(fam)
        if body:
            # YAML 来源：拼接 header + family 块
            merged_family[fam] = header_tpl + "\n" + body
        else:
            # 回退到代码常量（已含 header）
            merged_family[fam] = fallback_family[fam]

    return {
        "system_prompt": sys_tpl,
        "user_prompt_header": header_tpl,
        "family_templates": merged_family,
    }


def _build_schema_text(config: Dict) -> str:
    lines = []
    lines.append("# F 轴：金融业务场景（可多选，用分号分隔）")
    for f in config.get("fin_scene_schema", []):
        lines.append(f"- {f['code']} {f['name']}：{f['description']}")
    lines.append("\n# T 轴：任务类型（可多选，用分号分隔）")
    for t in config.get("task_type_schema", []):
        lines.append(f"- {t['code']} {t['name']}：{t['description']}")
    lines.append("\n# RVEC 标签体系")
    lines.append("## R/V/E（负向问题标签，可多选，每个标签需标注 P 等级）")
    for gk, tags in config.get("rvec_schema", {}).items():
        if gk.startswith("C"):
            continue
        lines.append(f"\n### {gk}")
        for tag in tags:
            lines.append(f"  - {tag['code']} {tag['name']}（{tag['type']}）")
    lines.append("\n## C（正向亮点标签，可多选，不抵消 R/V/E）")
    for gk, tags in config.get("rvec_schema", {}).items():
        if not gk.startswith("C"):
            continue
        lines.append(f"\n### {gk}")
        for tag in tags:
            lines.append(f"  - {tag['code']} {tag['name']}（{tag['type']}）")
    lines.append("\n# P 等级")
    sev = config.get("severity_schema", {})
    for lv in sev.get("levels", []):
        lines.append(f"- {lv}：{sev.get('definitions', {}).get(lv, '')}")
    scoring = config.get("scoring_rules", {})
    if scoring:
        lines.append(f"\n# 评分规则（0-4 分）")
        lines.append(scoring.get("description", ""))
    return "\n".join(lines)


def _build_few_shot_text(config: Dict) -> str:
    examples = config.get("few_shot_examples", [])
    if not examples:
        return ""
    lines = ["\n## 参考样例（请严格对齐以下标注风格和粒度）"]
    for i, ex in enumerate(examples, 1):
        labels = ex.get("expected_labels", {})
        lines.append(f"\n### 样例 {i}")
        lines.append(f"【题目】{ex.get('question', '')[:500]}")
        lines.append(f"【参考答案】{ex.get('expected_output', '')[:500]}")
        lines.append(f"【模型回答】{ex.get('answer', '')[:500]}")
        lines.append(f"【正确标注】\n```json\n{json.dumps(labels, ensure_ascii=False, indent=2)}\n```")
    return "\n".join(lines)


SYSTEM_PROMPT_TEMPLATE = """\
你是一位金融领域大模型评测专家，精通 RVEC 评测标签体系。你需要对模型在金融场景下的输出进行结构化评测。

## 评测流程
1. **识别需求侧画像**：判断金融业务场景（F轴）和任务类型（T轴）
2. **打 RVEC 标签**：识别 R（硬伤）、V（价值）、E（体验）负向问题 + C（亮点）
3. **判定 P 等级**：每个负向标签标注 P0/P1/P2，**严重度顺序为 P0 > P1 > P2（P0 最严重）**
4. **计算最终评分**：根据最严重负向信号给出 0-4 分

## 核心原则
- R 优先：R-P0 直接判 0 分，不因 V/E/C 好而抵消
- C 不抵消：亮点不能抵消 R/V/E 的负向问题
- P2 二次判断：仅 P2 时需判断是否需要修改（需改=2分，不需改=3分）
- 单一问题最小归因：优先选最小最直接的标签

## 标签体系
{schema_text}
{few_shot_text}

## 输出要求
你必须以**严格 JSON**格式输出，不要包含任何其他文字。格式如下：
{{
  "label_fin_scene": "F08 金融风险...",
  "label_task_type": "T16 题目作答...；T04 计算、测算...",
  "label_rve_primary": "R-REA-2 推理跳步",
  "label_rve_all": "R-REA-2 推理跳步：P1；R-FACT-2 计算错误：P2",
  "label_rve_score_all": "R-REA-2:0.7;R-FACT-2:0.3",
  "label_severity": "P1",
  "label_score": 1,
  "label_highlights": "C-R-3 逻辑完整",
  "label_evidence": "摘录有问题的具体句子",
  "label_reason": "判定理由"
}}

字段说明：
- label_fin_scene：金融场景标签（F轴），多选用「；」分隔，首个为 primary
- label_task_type：任务类型标签（T轴），多选用「；」分隔，首个为 primary。T16 必须搭配实际任务 T 标签
- label_rve_primary：最主要的一个 RVE 负向标签（仅 R/V/E 三轴，不含 C），无问题填 "NONE"
- label_rve_all：全部 RVE 负向标签，「；」分隔，**每个标签末尾用「：P0/P1/P2」内嵌严重程度**（如 "R-FACT-1 事实错误：P1；R-REA-4 前提错误：P0"），无问题填 "NONE"
- label_rve_score_all：**【必填字段】**各负向信号对最终判定的独立影响分值（0~1），格式 "标签编码:分值;标签编码:分值"（如 "R-FACT-1:0.8;R-REA-4:1.0"）。**只要 label_rve_all 不是 "NONE"，本字段就必须给出对应每个标签的分值，分值数量与顺序必须与 label_rve_all 完全对齐**。无问题时填 "NONE"。
- label_severity：最严重的 P 等级（P0/P1/P2/NONE）。**注意：此字段会由脚本从 label_rve_all 自动派生覆盖，但你仍需输出**
- label_score：0-4 分整数（依据 scoring_rules）
- label_highlights：C 亮点标签（仅 C 轴），「；」分隔，无亮点填 "NONE"
- label_evidence：摘录有问题的关键句子（从模型回答或题面中），无问题留空
- label_reason：每个标签的判定理由
"""

USER_PROMPT_HEADER = """\
请对以下模型回答进行金融 RVEC 评测。

【题型族】{family_display}
【题型说明】{family_features}
【该题型重点关注的负向标签】{family_focus}
{family_extra_rules}
"""

USER_PROMPT_QA_CHOICE = USER_PROMPT_HEADER + """\
【题目/问题】
{question}

【完整题面/附件（模型作答时可见的完整上下文）】
{context}

【参考答案（含正确选项 + 解释）】
{expected_output}

【模型实际回答】
{answer}

【模型选项】{model_choice}
【期望选项】{expected_choice}

【模型推理过程】
{reasoning}

【原始 Accuracy 评分】{accuracy}
{judge_summary}
{truncation_warning}

## 重要标注规则提醒（CFA/金融选择题专用）
1. **考试题归因强制**：当模型答错（Accuracy<1）时，`label_rve_primary` **必须**是导致错误的过程标签（R-REA-2 推理跳步、R-REA-4 前提错误、R-REA-6 概念误用、R-FACT-2 计算错误、R-UND-2 理解偏差）。**禁止用 R-FACT-1 作为考试题主标签**——R-FACT-1 仅在模型对核心知识点本身就是错的（纯知识缺失、无推理过程可归因）时使用。
2. **T16 联合标注**：考试场景必须同时标注 T16 + 至少一个实际任务标签（T04 计算/T02 事实/T06 分析归因等），T16 不可单独出现。
3. **答对的题不标 R-FACT**：Accuracy=1 时，回答与参考一致，不应标 R-FACT-1/R-FACT-7。如有问题归 V/E 维度。
4. **完整题面/附件为"无附加题面信息"时**：审慎使用 R-FACT-7（信息编造）——模型作答时可能有我们看不到的背景材料。
"""

USER_PROMPT_SENTIMENT = USER_PROMPT_HEADER + """\
【新闻文本/题目】
{question}

【完整文本】
{context}

【期望情感标签】
{expected_output}

【模型实际回答】
{answer}

【模型推理过程】
{reasoning}

【原始 Accuracy 评分】{accuracy}
{judge_summary}
{truncation_warning}

## 重要标注规则提醒（金融新闻情感分析专用）
1. **情感判断错误的归因**：优先归 R-UND-2 题意误解 或 R-UND-1 主体识别错误，**不要轻易用 R-FACT-1**（情感判断不是事实查询）。
2. **空泛回答**：回答冗长但无实质内容 → V-INFO-1 内容空洞；空泛附和无证据 → V-INFO-4 证据不足。
3. **语气不匹配**：情感标签准确但语气与新闻调性不符 → E-CONS-2 受众跳变 或 E-ADAPT-2 风格/受众不匹配。
4. **答对的题不标 R-FACT**：Accuracy=1 时不应标 R-FACT-*。
"""

USER_PROMPT_REPORT_EVAL = USER_PROMPT_HEADER + """\
【任务说明/问题】
{question}

【完整背景材料（PDF 摘要或片段）】
{context}

【参考材料】
{ground_truth_unstructured}

【模型实际回答】
{answer}

【模型推理过程】
{reasoning}

【多维评分】
{multi_dim_scores}
{judge_summary}
{truncation_warning}

## 重要标注规则提醒（研报/材料分析专用）
1. **信息编造（R-FACT-7）的判定**：模型回答中的信息如果能从【完整背景材料】中找到来源，则不算编造；只有完全凭空捏造才标 R-FACT-7。**【完整背景材料】为空时不可轻判 R-FACT-7**。
2. **优先依据 judge_comment.factual_check**：若 factual_check 已标出错误事实点，对应归 R-FACT-7 编造 或 R-FACT-1 事实错误。
3. **优先依据 judge_comment.recall_check**：若 recall_check 标出遗漏的核心论点，归 R-FACT-3 漏答。
4. **截断警告**：若 model_response_truncated=yes，禁止仅因末尾缺失判 R-FACT-3 漏答。
5. **数据口径**：研报中的指标数字若单位/时点/口径错误 → R-FACT-8 数据口径错误。
6. **结构问题**：长报告关注 E-STR-1 层次混乱、E-STR-2 关键结论不突出。
"""

USER_PROMPT_LONG_GEN = USER_PROMPT_HEADER + """\
【任务说明/问题】
{question}

【完整背景材料】
{context}

【模型实际回答】
{answer}

【模型推理过程】
{reasoning}

【多维评分（无显式参考答案，请综合判断）】
{multi_dim_scores}
{judge_summary}
{truncation_warning}

## 重要标注规则提醒（长报告生成 / 无显式 GT 专用）
1. **无 GT 场景下的事实判断**：主要依据 judge_comment.factual_check；优先归 R-FACT-7 编造 或 R-FACT-8 数据口径。
2. **数据口径**：报告中的财务指标若单位/时点/口径错误 → R-FACT-8。
3. **内容质量**：空洞无信息量 → V-INFO-1；分析浅 → V-INFO-2 分析浅显；结论无支撑 → V-INFO-4 证据不足。
4. **亮点**：长报告若结构清晰、多方案对比、风险提示充分，可给 C-V-1 多方案对比、C-R-2 风险完整、C-E-2 结构友好。
5. **截断警告**：若 model_response_truncated=yes，禁止仅因末尾缺失判 R-FACT-3 漏答。
6. **若 comprehensive_score 较高（>0.7）**但仍有问题，倾向标 P2 而非 P1。
"""

USER_PROMPT_GENERIC = USER_PROMPT_HEADER + """\
【题目/问题】
{question}

【完整题面/附件】
{context}

【参考答案】
{expected_output}

【模型实际回答】
{answer}

【模型推理过程】
{reasoning}

【原始 Accuracy 评分】{accuracy}
{judge_summary}
{truncation_warning}

## 重要标注规则提醒（通用）
1. **信息编造判定**：模型回答中的信息如果能从【完整题面/附件】中找到来源，则不算编造（R-FACT-7）。
2. **T16 联合标注**：如果属于考试/作答场景，T16 必须搭配至少一个实际任务 T 标签。
3. **答对的题不标 R-FACT**：Accuracy=1 时不应标 R-FACT-* 类。
"""

FAMILY_TEMPLATES = {
    "QA_CHOICE": USER_PROMPT_QA_CHOICE,
    "SENTIMENT": USER_PROMPT_SENTIMENT,
    "REPORT_EVAL": USER_PROMPT_REPORT_EVAL,
    "LONG_GEN": USER_PROMPT_LONG_GEN,
    "GENERIC": USER_PROMPT_GENERIC,
}

# ════════════════════════════════════════════════════════════
# 4. 字段映射（适配 merge_stage1.py 新 schema）
# ════════════════════════════════════════════════════════════

_FIELD_CANDIDATES = {
    # 题目主体（去掉前置背景材料的核心提问）
    "question": ["question", "题目", "prompt", "query", "问题"],
    # 背景上下文（report 池含 PDF 摘要，QA 池可能为空）
    "context": ["context", "输入", "input", "full_context", "题干"],
    # 模型最终回答（已剥离 think/CoT）
    "answer": ["model_response", "实际回答", "output", "answer", "answer_text_for_labeling", "response", "回答"],
    # 选择题答案字母
    "model_choice": ["model_choice"],
    # 结构化标准答案（QA）
    "reference": ["ground_truth_structured", "参考答案", "expected_output", "reference", "gold", "标准答案"],
    # 非结构化标准答案（report 池 PDF 文件名）
    "ground_truth_unstructured": ["ground_truth_unstructured"],
    # 模型推理过程（JSON 或文本）
    "reasoning": ["model_reasoning", "推理过程", "reasoning", "trace_output", "思维链"],
    # 评分
    "accuracy": ["Accuracy", "accuracy", "score", "评分"],
    # 自动评测系统的评注（JSON）
    "judge_comment": ["judge_comment"],
    # 截断标记
    "truncated": ["model_response_truncated"],
    "full_length": ["model_response_full_length"],
    # 多维评分（report 池）
    "factuality_score": ["factuality_score"],
    "recall_score": ["recall_score"],
    "reasoning_score": ["reasoning_score"],
    "structure_score": ["structure_score"],
    "comprehensive_score": ["comprehensive_score"],
    # 已有的文字评注
    "accuracy_comment": ["Accuracy_comment", "accuracy_comment"],
    "reasoning_quality_comment": ["reasoning_quality_comment"],
    # 题型族识别
    "norm_task": ["norm_task_from_filename"],
    "meta_task_family": ["meta_task_family"],
    "meta_schema_type": ["meta_schema_type"],
    # ID
    "id": ["answer_id", "id", "item_id", "sample_id"],
}


def _map_field(record: Dict, role: str) -> str:
    for col in _FIELD_CANDIDATES.get(role, []):
        val = record.get(col, "")
        if val and str(val).strip():
            return str(val)
    return ""


def _auto_field_mapping(records: List[Dict]) -> Dict[str, str]:
    if not records:
        return {}
    columns = list(records[0].keys())
    mapping = {}
    for role, candidates in _FIELD_CANDIDATES.items():
        for col in candidates:
            if col in columns:
                mapping[role] = col
                break
    return mapping


# ════════════════════════════════════════════════════════════
# 5. 打标核心逻辑
# ════════════════════════════════════════════════════════════

# 用于从 label_rve_all 中解析 P 等级
_P_LEVEL_RE = re.compile(r'P[012]')

# 严重度排序：P0 最严重 > P1 > P2 > NONE
_SEVERITY_ORDER = {"P0": 3, "P1": 2, "P2": 1, "NONE": 0, "": 0}


def _parse_json_field(val: str) -> str:
    """解析可能是 JSON 字符串的字段，提取文本内容。
    适配 model_reasoning 字段（JSON: {reasoning_text, trace_raw}）和旧 CFA dict 格式。
    """
    if not val:
        return ""
    val = str(val).strip()
    if not val:
        return ""
    # JSON 解析
    if val.startswith("{"):
        try:
            parsed = json.loads(val)
            if isinstance(parsed, dict):
                # 优先 reasoning_text（model_reasoning 新格式）
                if "reasoning_text" in parsed:
                    rt = str(parsed.get("reasoning_text", "")).strip()
                    if rt:
                        return rt
                    # reasoning_text 为空，尝试 trace_raw
                    tr = parsed.get("trace_raw", "")
                    if tr:
                        if isinstance(tr, (dict, list)):
                            return json.dumps(tr, ensure_ascii=False)[:3000]
                        return str(tr)[:3000]
                    return ""
                # 其他常见字段
                for k in ("content", "text", "reasoning", "reasoning_process", "model_cot", "choice", "model_choice"):
                    if k in parsed and parsed[k]:
                        return str(parsed[k])
                return val
        except (json.JSONDecodeError, TypeError):
            pass
    return val


def _derive_severity(rve_all: str) -> str:
    """从 label_rve_all 解析所有 P 等级，返回最严重的（P0 > P1 > P2）。
    若无 P 等级，返回 NONE。
    """
    if not rve_all or rve_all.strip() in ("", "NONE"):
        return "NONE"
    levels = _P_LEVEL_RE.findall(str(rve_all))
    if not levels:
        return "NONE"
    # P0 > P1 > P2
    if "P0" in levels:
        return "P0"
    if "P1" in levels:
        return "P1"
    if "P2" in levels:
        return "P2"
    return "NONE"


def _validate_score(severity: str, rve_all: str, highlights: str, raw_score) -> int:
    """根据 scoring_rules 校验最终评分。"""
    try:
        raw_score = int(raw_score)
    except (ValueError, TypeError):
        raw_score = 3
    if severity == "P0":
        return 0
    elif severity == "P1":
        return 1
    elif severity == "P2":
        return min(raw_score, 3) if raw_score >= 2 else 2
    elif rve_all == "NONE" or not rve_all.strip():
        return 4 if (highlights != "NONE" and highlights.strip()) else 3
    return raw_score


def classify_family(record: Dict, config: Dict) -> str:
    """根据 record 的 norm_task_from_filename 等字段，识别题型族。
    返回 family key（QA_CHOICE / SENTIMENT / REPORT_EVAL / LONG_GEN / GENERIC）。
    """
    rules = config.get("task_family_rules", {})
    norm_task = str(_map_field(record, "norm_task")).strip()
    if not norm_task:
        return "GENERIC"
    for family, cfg in rules.items():
        if family.startswith("_"):
            continue
        match_list = cfg.get("match_norm_task", [])
        if norm_task in match_list:
            return family
    return "GENERIC"


def _summarize_judge_comment(judge_comment_str: str, family: str, max_len: int = 1500) -> str:
    """把 judge_comment（JSON）摘要后注入到 prompt。
    QA 池含: accuracy_comment / expected_choice / actual_choice / reasoning_quality_comment
    Report 池含: factual_check / recall_check / grounding_check / coverage_check / summary / reasoning_chain_quality
    实际上同一份 norm_task 的 judge_comment 字段可能混合（例如 Eval_FullReport 的 judge_comment 实为 QA 风格），
    因此采用"按 family 主字段优先 + 跨 family 兜底"策略，避免漏掉关键线索。
    """
    if not judge_comment_str:
        return ""
    try:
        jc = json.loads(judge_comment_str) if isinstance(judge_comment_str, str) else judge_comment_str
    except (json.JSONDecodeError, TypeError):
        return ""
    if not isinstance(jc, dict):
        return ""

    # 两套字段集
    QA_FIELDS = [
        ("accuracy_comment", "评分说明"),
        ("expected_choice", "期望选项"),
        ("actual_choice", "实际选项"),
        ("reasoning_quality_comment", "推理质量评注"),
    ]
    REPORT_FIELDS = [
        ("summary", "评测总结"),
        ("factual_check", "事实核查"),
        ("recall_check", "召回核查"),
        ("grounding_check", "依据核查"),
        ("coverage_check", "覆盖核查"),
        ("reasoning_chain_quality", "推理链质量"),
    ]

    # 按 family 决定优先级
    if family in ("REPORT_EVAL", "LONG_GEN"):
        ordered = REPORT_FIELDS + QA_FIELDS  # 先 report，后 QA 兜底
        per_field_limit = 300
    else:
        ordered = QA_FIELDS + REPORT_FIELDS  # 先 QA，后 report 兜底
        per_field_limit = 400

    lines = []
    seen_keys = set()
    for k, label in ordered:
        if k in seen_keys:
            continue
        seen_keys.add(k)
        v = jc.get(k)
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        if isinstance(v, (dict, list)):
            vs = json.dumps(v, ensure_ascii=False)
        else:
            vs = str(v).strip()
        if not vs:
            continue
        lines.append(f"- {label}：{vs[:per_field_limit]}")

    if not lines:
        return ""
    body = "\n".join(lines)
    if len(body) > max_len:
        body = body[:max_len] + "...(自动评测摘要已截断)"
    return f"\n【自动评测系统线索（judge_comment 摘要）】\n{body}\n"


def _build_truncation_warning(record: Dict) -> str:
    """若 model_response_truncated=yes，输出警告段。"""
    truncated = str(_map_field(record, "truncated")).strip().lower()
    if truncated == "yes":
        full_len = _map_field(record, "full_length") or "未知"
        return (
            f"\n⚠️【截断警告】此模型回答已被截断（原始完整长度 {full_len} 字符，"
            f"当前可见部分约 {len(_map_field(record, 'answer'))} 字符）。"
            f"请勿仅因末尾信息缺失而判 R-FACT-3 漏答。\n"
        )
    return ""


def _build_multi_dim_scores(record: Dict) -> str:
    """组装 report 池的多维评分文本。"""
    parts = []
    for role, label in [
        ("comprehensive_score", "综合得分"),
        ("factuality_score", "事实性"),
        ("recall_score", "召回率"),
        ("reasoning_score", "推理"),
        ("structure_score", "结构"),
    ]:
        v = _map_field(record, role)
        if v and str(v).strip():
            parts.append(f"{label}={v}")
    return " | ".join(parts) if parts else "（无多维评分）"


def _build_family_header(family: str, config: Dict) -> Dict[str, str]:
    """构造 family header 三段：display_name / input_features / focus_labels / extra_rules"""
    rules = config.get("task_family_rules", {}).get(family, {})
    extra = rules.get("extra_rules", [])
    extra_text = ""
    if extra:
        extra_text = "【该题型特别规则】\n" + "\n".join(f"  - {r}" for r in extra)
    return {
        "family_display": rules.get("display_name", family),
        "family_features": rules.get("input_features", "（无）"),
        "family_focus": "、".join(rules.get("focus_labels", [])) or "（无特定限制）",
        "family_extra_rules": extra_text,
    }


# 解析 rve_all 单标签：(整段, 标签编码, 标签名, P等级)
_TAG_SPLIT_RE = re.compile(r'^(.+?)：(P[012])\s*$')
# 提取标签编码（开头到第一个空格）
_CODE_RE = re.compile(r'^([A-Z]+(?:-[A-Z0-9]+)+)')


def _parse_rve_tag(tag_str: str) -> Optional[Dict[str, str]]:
    """解析单个 rve 标签字符串 'R-FACT-1 事实错误：P1' → {'full':..., 'code':..., 'name':..., 'p_level':...}"""
    tag_str = tag_str.strip()
    if not tag_str or tag_str == "NONE":
        return None
    m = _TAG_SPLIT_RE.match(tag_str)
    if m:
        body = m.group(1).strip()
        p_level = m.group(2)
    else:
        body = tag_str
        p_level = ""
    cm = _CODE_RE.match(body)
    code = cm.group(1) if cm else body.split(" ")[0]
    name = body[len(code):].strip()
    return {"full": tag_str, "body": body, "code": code, "name": name, "p_level": p_level}


def _parse_score_all(score_all_str: str) -> Dict[str, str]:
    """解析 label_rve_score_all 'R-FACT-1:0.9;R-REA-2:0.4' → {'R-FACT-1':'0.9', ...}"""
    result = {}
    if not score_all_str or score_all_str.strip() in ("", "NONE"):
        return result
    # 兼容中英文分隔符
    for piece in re.split(r'[;；]', str(score_all_str)):
        piece = piece.strip()
        if not piece:
            continue
        if ":" in piece:
            k, v = piece.split(":", 1)
            result[k.strip()] = v.strip()
    return result


def _post_process_labels(result: Dict, accuracy: str) -> Dict:
    """统一后处理（LLM 输出兜底校验 + 重排 + 派生）。
    顺序很关键：
      1. 解析 rve_all → 标签列表
      2. 答对的题（Accuracy=1）剔除所有 R-FACT-* 标签
      3. 考试题（T16）且 primary=R-FACT-1 → 改为过程标签
      4. 确保 primary 在 rve_all 列表的第一个（重排）
      5. 同步重排 rve_score_all
      6. 派生 severity（P0 > P1 > P2）
      7. 校验 label_score
    """
    task_type = str(result.get("label_task_type", ""))
    rve_all_str = str(result.get("label_rve_all", "")).strip()
    primary_str = str(result.get("label_rve_primary", "")).strip()
    score_all_str = str(result.get("label_rve_score_all", "")).strip()
    highlights = str(result.get("label_highlights", "NONE")).strip()

    # ── 1. 解析 rve_all 标签列表 ──
    tags = []
    if rve_all_str and rve_all_str != "NONE":
        for piece in re.split(r'[;；]', rve_all_str):
            parsed = _parse_rve_tag(piece)
            if parsed:
                tags.append(parsed)

    # 解析独立分值映射
    score_map = _parse_score_all(score_all_str)

    # ── 2. 答对的题强制移除 R-FACT-* ──
    if str(accuracy).strip() in ("1", "1.0"):
        before_n = len(tags)
        tags = [t for t in tags if not t["code"].startswith("R-FACT")]
        # 同步移除 score_map 中的 R-FACT-*
        score_map = {k: v for k, v in score_map.items() if not k.startswith("R-FACT")}
        if len(tags) < before_n:
            logger.debug(f"答对题剔除 R-FACT-* 标签 {before_n - len(tags)} 个")
        # primary 若是 R-FACT-* 也清掉
        if primary_str.startswith("R-FACT") or "R-FACT" in primary_str.split(" ")[0]:
            primary_str = ""

    # ── 3. 考试题（T16）且 primary=R-FACT-1 → 改为过程标签 ──
    if "T16" in task_type and primary_str:
        primary_code = primary_str.split(" ")[0] if " " in primary_str else primary_str
        if primary_code == "R-FACT-1":
            # 从 tags 中找过程标签作为新 primary
            for t in tags:
                if (t["code"].startswith("R-REA") or t["code"].startswith("R-UND")
                        or t["code"] == "R-FACT-2"):
                    primary_str = t["body"]  # 不带 P 等级的版本
                    break

    # ── 4. 确保 primary 在 rve_all 列表第一个 ──
    if primary_str and tags:
        primary_code = primary_str.split(" ")[0] if " " in primary_str else primary_str
        # 找到 primary 在 tags 中的索引
        primary_idx = -1
        for i, t in enumerate(tags):
            if t["code"] == primary_code:
                primary_idx = i
                break
        if primary_idx > 0:
            # 把 primary 移到第一位
            tags = [tags[primary_idx]] + tags[:primary_idx] + tags[primary_idx + 1:]
        elif primary_idx == -1 and tags:
            # primary 不在 rve_all 列表中：用 rve_all 第一个反推 primary
            primary_str = tags[0]["body"]

    # ── 5. 重组 rve_all 和 rve_score_all（保持顺序对齐）──
    # P 等级 → 默认分值（兜底用，当 LLM 没给 score_all 时按 P 等级派生）
    P_DEFAULT_SCORE = {"P0": 1.0, "P1": 0.7, "P2": 0.3, "": 0.5}
    if tags:
        result["label_rve_all"] = "；".join(t["full"] for t in tags)
        # score_all 按 tags 顺序输出；缺失分值时按 P 等级派生（primary 额外 +0.1，封顶 1.0）
        new_score_pairs = []
        for i, t in enumerate(tags):
            if t["code"] in score_map:
                # LLM 已给出，原样保留
                new_score_pairs.append(f"{t['code']}:{score_map[t['code']]}")
            else:
                # 兜底：按 P 等级派生
                base = P_DEFAULT_SCORE.get(t["p_level"], 0.5)
                if i == 0:  # primary（第一位）权重略高
                    base = min(base + 0.1, 1.0)
                # 保留 1 位小数（避免 0.7000000001 这种）
                new_score_pairs.append(f"{t['code']}:{round(base, 2)}")
        result["label_rve_score_all"] = ";".join(new_score_pairs)
        result["label_rve_primary"] = primary_str or tags[0]["body"]
    else:
        # 无 RVE 负向标签
        result["label_rve_all"] = "NONE"
        result["label_rve_score_all"] = "NONE"
        result["label_rve_primary"] = "NONE"

    # ── 6. 派生 severity（P0 > P1 > P2，脚本覆盖 LLM 输出）──
    new_rve_all = str(result.get("label_rve_all", "NONE"))
    derived_severity = _derive_severity(new_rve_all)
    result["label_severity"] = derived_severity

    # ── 7. 校验 label_score ──
    result["label_score"] = _validate_score(
        derived_severity,
        new_rve_all,
        highlights,
        result.get("label_score", 3),
    )

    return result


def label_one(record: Dict, llm: LLMClient, system_prompt: str, config: Dict) -> Dict:
    # ── 1. 抽字段 ──
    question = _parse_json_field(_map_field(record, "question"))
    context = _parse_json_field(_map_field(record, "context"))
    answer = _parse_json_field(_map_field(record, "answer"))
    reference = _map_field(record, "reference")
    ground_truth_unstructured = _map_field(record, "ground_truth_unstructured")
    reasoning = _parse_json_field(_map_field(record, "reasoning"))
    accuracy = _map_field(record, "accuracy") or "（无）"
    model_choice = _map_field(record, "model_choice") or "（无）"
    judge_comment = _map_field(record, "judge_comment")

    # 期望选项：从 judge_comment 提取
    expected_choice = "（无）"
    try:
        jc = json.loads(judge_comment) if judge_comment else {}
        if isinstance(jc, dict):
            ec = jc.get("expected_choice", "")
            if ec:
                expected_choice = str(ec)
    except (json.JSONDecodeError, TypeError):
        pass

    # context 和 question 相同时清空，避免重复
    if context and context == question:
        context = ""

    # ── 2. 识别 family + 选模板 ──
    family = classify_family(record, config)
    # 优先用 config["_family_templates"]（由 _resolve_prompts 注入），未注入则回退到全局常量
    family_templates = config.get("_family_templates") or FAMILY_TEMPLATES
    template = family_templates.get(family) or family_templates.get("GENERIC") or USER_PROMPT_GENERIC
    header = _build_family_header(family, config)
    judge_summary = _summarize_judge_comment(judge_comment, family)
    truncation_warning = _build_truncation_warning(record)
    multi_dim_scores = _build_multi_dim_scores(record)

    # ── 3. 组 USER_PROMPT ──
    # 兼容所有模板都用到的占位符（即使该模板不用，也要填值避免 KeyError）
    fmt_kwargs = dict(
        family_display=header["family_display"],
        family_features=header["family_features"],
        family_focus=header["family_focus"],
        family_extra_rules=header["family_extra_rules"],
        question=question[:3000],
        context=context[:4000] if context else "（无附加题面信息）",
        expected_output=reference[:2000] if reference else "（无）",
        ground_truth_unstructured=ground_truth_unstructured or "（无）",
        answer=answer[:5000],
        reasoning=reasoning[:3000] if reasoning else "（无）",
        accuracy=accuracy,
        model_choice=model_choice,
        expected_choice=expected_choice,
        judge_summary=judge_summary,
        truncation_warning=truncation_warning,
        multi_dim_scores=multi_dim_scores,
    )
    user_msg = template.format(**fmt_kwargs)

    # ── 4. 调 LLM ──
    labels = llm.chat_json(system_prompt, user_msg)

    # ── 5. F×T 交叉校验 + 二次修正 ──
    labels["_accuracy_hint"] = accuracy
    labels["_family_hint"] = family
    ft_hint = _get_ft_cross_hint(config, labels)
    labels.pop("_accuracy_hint", None)
    labels.pop("_family_hint", None)
    if ft_hint:
        verify_msg = (
            f"你刚才给出的标注结果如下：\n```json\n{json.dumps(labels, ensure_ascii=False, indent=2)}\n```\n"
            f"{ft_hint}\n\n请根据交叉校验提醒重新审视标注，输出修正后的完整JSON。"
        )
        try:
            labels = llm.chat_json(system_prompt, verify_msg)
        except Exception:
            pass

    # ── 6. 写回结果（用新字段名 label_rve_*） ──
    result = dict(record)
    for key in ["label_fin_scene", "label_task_type",
                "label_rve_primary", "label_rve_all", "label_rve_score_all",
                "label_severity", "label_score",
                "label_highlights", "label_evidence", "label_reason"]:
        result[key] = labels.get(key, "")

    # ── 7. 后处理（一站式：答对剔 R-FACT / 主标签纠正 / rve_all 重排 / score 兜底）──
    result = _post_process_labels(result, accuracy)

    # ── 8. labeler 标识：fin_rvec_tag@<model 名>（动态记录所用 LLM）──
    model_tag = str(getattr(llm, "model", "") or "unknown").strip().replace(" ", "_")
    result["labeler"] = f"fin_rvec_tag@{model_tag}"
    result["review_status"] = result.get("review_status", "") or "pending"

    return result


def _get_ft_cross_hint(config: Dict, labels: Dict) -> str:
    scene_codes = re.findall(r'F\d+', str(labels.get("label_fin_scene", "")))
    task_codes = re.findall(r'T\d+', str(labels.get("label_task_type", "")))
    if not scene_codes or not task_codes:
        return ""
    hints = []

    # 硬约束：T16 不可单独出现
    if "T16" in task_codes and len(task_codes) == 1:
        hints.append("⚠️ T16（题目作答）不可单独出现，必须搭配至少一个描述实际任务内容的 T 标签（如 T04 计算、T02 事实核验、T06 分析归因等）。请补充实际任务标签。")

    # 考试题归因检查：如果有 T16 且主标签是 R-FACT-1，强制要求修改
    primary = str(labels.get("label_rve_primary", ""))
    if "T16" in task_codes:
        if "R-FACT-1" in primary:
            hints.append("❌ 强制修正：考试题场景下，禁止使用 R-FACT-1（事实错误）作为主标签。必须改为导致错误的过程标签：R-REA-4（前提错误）、R-REA-6（概念误用）、R-REA-2（推理跳步）、R-FACT-2（计算错误）、R-UND-2（理解偏差）等。R-FACT-1 仅在纯知识缺失（模型不知道这个知识点）且无推理过程可归因时使用。请修改 label_rve_primary。")

    # 答对了却标了 R-FACT 类标签：强制纠正
    accuracy_val = str(labels.get("_accuracy_hint", ""))
    rve_all = str(labels.get("label_rve_all", ""))
    if accuracy_val in ("1", "1.0") and ("R-FACT" in rve_all or "R-FACT" in primary):
        hints.append("❌ 强制修正：原始 Accuracy=1 表示模型答案正确，不应标注 R-FACT 类标签。如有问题请改为 V 或 E 维度（如 V-INFO-4 证据不足）。")

    # family 专属强制规则
    family = str(labels.get("_family_hint", ""))
    family_rules = config.get("task_family_rules", {}).get(family, {})
    forbid_primary = family_rules.get("forbid_primary", [])
    for forbid in forbid_primary:
        if forbid in primary:
            hints.append(f"❌ 强制修正：{family} 题型族禁止用 {forbid} 作为 primary 标签，请改选其他过程标签。")

    rules = config.get("ft_cross_validation", {}).get("rules", [])
    for rule in rules:
        if (any(s in rule.get("scenes", []) for s in scene_codes) and
                any(t in rule.get("tasks", []) for t in task_codes)):
            hints.append(f"⚠️ {rule['note']}（必检: {', '.join(rule['must_check'])}）")
    return "\n\n【F×T交叉校验】\n" + "\n".join(hints) if hints else ""


# ════════════════════════════════════════════════════════════
# 6. 批量执行（并发 + 进度 + 重试）
# ════════════════════════════════════════════════════════════

def run_batch(records: List[Dict], llm: LLMClient, system_prompt: str, config: Dict,
              workers: int = 5, retry_limit: int = 3, progress_file: Optional[str] = None) -> tuple:
    success, failed = [], []
    total = len(records)
    done_count = 0
    fail_count = 0
    start_time = time.time()
    lock = threading.Lock()

    def _update_progress():
        """更新进度文件（JSON），供外部轮询"""
        elapsed = time.time() - start_time
        speed = done_count / max(elapsed, 0.01)
        remaining = (total - done_count) / max(speed, 0.01) if speed > 0 else 0
        progress = {
            "status": "running",
            "total": total,
            "done": done_count,
            "success": done_count - fail_count,
            "failed": fail_count,
            "percent": round(done_count / total * 100, 1),
            "speed": round(speed, 2),
            "elapsed_s": round(elapsed, 1),
            "remaining_s": round(remaining, 1),
            "eta": datetime.now().replace(microsecond=0).isoformat() if remaining < 1 else
                   (datetime.now() + timedelta(seconds=remaining)).replace(microsecond=0).isoformat(),
            "updated_at": datetime.now().isoformat(),
        }
        if progress_file:
            try:
                Path(progress_file).write_text(json.dumps(progress, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception:
                pass
        return progress

    def _do_one(record):
        nonlocal done_count, fail_count
        time.sleep(random.uniform(0.3, 1.0))  # 请求间隔，降低限流概率
        result = label_one(record, llm, system_prompt, config)
        result["label_status"] = "done"
        result["labeled_at"] = datetime.now().isoformat()
        with lock:
            done_count += 1
            p = _update_progress()
            # 每条打印一行结构化进度（Agent 可解析）
            print(f"\r⚡ RVEC打标: {p['done']}/{p['total']} ({p['percent']}%) | "
                  f"✅{p['success']} ❌{p['failed']} | "
                  f"{p['speed']:.1f}条/s | 剩余~{p['remaining_s']:.0f}s", end="", flush=True)
        return result

    if workers > 1:
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {pool.submit(_do_one, r): r for r in records}
            for fut in concurrent.futures.as_completed(futures):
                rec = futures[fut]
                try:
                    success.append(fut.result())
                except Exception as e:
                    with lock:
                        done_count += 1
                        fail_count += 1
                        _update_progress()
                    logger.warning(f"❌ 打标失败: {str(e)[:120]}")
                    failed.append(rec)
    else:
        for r in records:
            try:
                success.append(_do_one(r))
            except Exception as e:
                with lock:
                    done_count += 1
                    fail_count += 1
                    _update_progress()
                logger.warning(f"❌ 打标失败: {str(e)[:120]}")
                failed.append(r)
    print()  # 换行

    # 重试
    for attempt in range(1, retry_limit + 1):
        if not failed:
            break
        logger.info(f"🔄 重试 {attempt}/{retry_limit}，剩余失败: {len(failed)}")
        still_failed = []
        for r in failed:
            try:
                result = label_one(r, llm, system_prompt, config)
                result["label_status"] = "done"
                result["labeled_at"] = datetime.now().isoformat()
                success.append(result)
            except Exception:
                still_failed.append(r)
        failed = still_failed

    for r in failed:
        r["label_status"] = "failed"

    # 写入最终进度
    elapsed = time.time() - start_time
    final_progress = {
        "status": "completed",
        "total": total,
        "done": total,
        "success": len(success),
        "failed": len(failed),
        "percent": 100.0,
        "elapsed_s": round(elapsed, 1),
        "speed": round(total / max(elapsed, 0.01), 2),
        "remaining_s": 0,
        "updated_at": datetime.now().isoformat(),
    }
    if progress_file:
        try:
            Path(progress_file).write_text(json.dumps(final_progress, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass

    return success, failed


# ════════════════════════════════════════════════════════════
# 6.5 工作流子命令（show-rules / show-prompt / A-B 对比）
# ════════════════════════════════════════════════════════════

# 输入快照字段顺序与展示长度（共用配置）
_SNAPSHOT_FIELDS = [
    # role,                label,            max_len
    ("question",            "❓ 题目",         600),
    ("context",             "📄 上下文",       800),
    ("reference",           "✅ 参考答案",     500),
    ("ground_truth_unstructured", "📎 非结构化参考", 200),
    ("answer",              "🤖 模型回答",     1500),
    ("model_choice",        "🅰  模型选项",     50),
    ("reasoning",           "🧠 模型推理",     500),
    ("accuracy",            "📊 原始 Accuracy", 50),
]


def _build_input_snapshot(record: Dict, family: str = "") -> Dict[str, str]:
    """从 record 抽取关键输入字段的完整快照（供预览展示）。

    返回 OrderedDict 风格的 {label: value}，按 _SNAPSHOT_FIELDS 顺序，
    以及自动衍生的 family / judge_summary / truncation / multi_dim_scores 段。
    """
    snap = {}
    snap["__family__"] = family or "GENERIC"

    for role, label, max_len in _SNAPSHOT_FIELDS:
        raw = _map_field(record, role)
        if not raw:
            continue
        if role in ("question", "context", "answer", "reasoning"):
            raw = _parse_json_field(raw)
        s = str(raw).strip()
        if not s:
            continue
        if len(s) > max_len:
            snap[label] = s[:max_len] + f"  ...(原始 {len(s)} 字符,已截断)"
        else:
            snap[label] = s

    # 期望选项（从 judge_comment 提取）
    judge_comment = _map_field(record, "judge_comment")
    if judge_comment:
        try:
            jc = json.loads(judge_comment)
            if isinstance(jc, dict):
                ec = jc.get("expected_choice", "")
                if ec:
                    snap["🎯 期望选项"] = str(ec)
        except (json.JSONDecodeError, TypeError):
            pass

    # judge_comment 摘要
    if judge_comment:
        js = _summarize_judge_comment(judge_comment, family)
        if js.strip():
            # _summarize_judge_comment 输出形如 "\n【自动评测系统线索（judge_comment 摘要）】\n- 评分说明：...\n"
            # 这里去掉头部的【...】标题行，只保留正文
            body_lines = []
            for ln in js.strip().splitlines():
                ln = ln.strip()
                if not ln:
                    continue
                if ln.startswith("【") and ln.endswith("】"):
                    continue
                body_lines.append(ln)
            body = " | ".join(body_lines)
            if body:
                snap["💡 judge 线索"] = body[:600] + ("..." if len(body) > 600 else "")

    # 截断信号
    truncated = str(_map_field(record, "truncated")).strip().lower()
    if truncated == "yes":
        full_len = _map_field(record, "full_length") or "未知"
        snap["⚠️ 截断"] = f"yes（原始长度 {full_len}）"

    # 多维评分（report 池）
    multi = _build_multi_dim_scores(record)
    if multi and multi != "（无多维评分）":
        snap["📈 多维评分"] = multi

    # norm_task
    norm_task = _map_field(record, "norm_task")
    if norm_task:
        snap["🏷  norm_task"] = norm_task

    # ID
    rid = _map_field(record, "id")
    if rid:
        snap["🆔 ID"] = rid

    return snap


def _print_snapshot(snap: Dict[str, str], indent: str = "  ") -> None:
    """以多行、自动换行的方式打印输入快照（终端友好）。"""
    family = snap.pop("__family__", "")
    if family:
        print(f"{indent}🎬 题型族: {family}")
    for label, value in snap.items():
        # 长字段独占多行：第一行带 label，后续行缩进
        if len(value) > 80 or "\n" in value:
            print(f"{indent}{label}:")
            for line in value.splitlines() or [value]:
                # 每 100 字符再硬换行一次（避免终端横向滚动）
                while len(line) > 100:
                    print(f"{indent}    {line[:100]}")
                    line = line[100:]
                print(f"{indent}    {line}")
        else:
            print(f"{indent}{label}: {value}")


def _print_label_result(item: Dict, indent: str = "  ") -> None:
    """统一格式化打印打标结果（共用）。"""
    print(f"{indent}🏷  场景:    {item.get('label_fin_scene', '')}")
    print(f"{indent}🏷  任务:    {item.get('label_task_type', '')}")
    print(f"{indent}⚠️  主问题:  {item.get('label_rve_primary', '')}")
    rve_all = item.get('label_rve_all', '')
    if rve_all and rve_all != item.get('label_rve_primary', ''):
        print(f"{indent}⚠️  全部:    {rve_all}")
    score_all = item.get('label_rve_score_all', '')
    if score_all and score_all not in ('NONE', '', None):
        print(f"{indent}📊 各信号:  {score_all}")
    print(f"{indent}🔴 严重度:  {item.get('label_severity', '')}    ⭐ 评分: {item.get('label_score', '')}")
    hl = item.get('label_highlights', '')
    if hl and hl != 'NONE':
        print(f"{indent}✨ 亮点:    {hl}")
    ev = str(item.get('label_evidence', '')).strip()
    if ev:
        print(f"{indent}📌 证据:    {ev[:200]}")
    rs = str(item.get('label_reason', '')).strip()
    if rs:
        # 理由通常较长，独立一段
        print(f"{indent}💬 理由:")
        for line in rs.splitlines():
            while len(line) > 100:
                print(f"{indent}    {line[:100]}")
                line = line[100:]
            print(f"{indent}    {line}")


def _cmd_show_rules(config: Dict, section: str = "all", family: Optional[str] = None) -> None:
    """打印当前生效的规则（不调用 LLM）。

    section 取值：
      all       — 全部章节
      rvec      — RVEC 标签体系
      scoring   — 评分规则
      severity  — P 等级
      scenes    — F 轴
      tasks     — T 轴
      ft_cross  — F×T 交叉校验
      families  — 题型族路由
    """
    # 数据来源标记（YAML / 内置）
    yaml_loaded = bool(config.get("_prompts"))
    print(f"\n{'='*70}")
    print(f"📜 当前生效的 RVEC 规则")
    print(f"   配置来源: {'config/rvec/*.yaml（已加载）' if yaml_loaded else 'fin_rvec_config（base）'}")
    if family:
        print(f"   过滤 family: {family}")
    print(f"{'='*70}")

    show = section.lower()
    show_all = (show == "all")

    if show_all or show == "scenes":
        print("\n## F 轴：金融业务场景")
        for f in config.get("fin_scene_schema", []):
            print(f"  {f['code']:6s} {f['name']}")
            if f.get("description"):
                print(f"         └─ {f['description']}")

    if show_all or show == "tasks":
        print("\n## T 轴：任务类型")
        for t in config.get("task_type_schema", []):
            print(f"  {t['code']:6s} {t['name']}")

    if show_all or show == "rvec":
        print("\n## RVEC 标签体系")
        for gk, tags in config.get("rvec_schema", {}).items():
            kind = "亮点（C）" if gk.startswith("C") else "负向（R/V/E）"
            print(f"\n  ### {gk}  [{kind}]")
            for tag in tags:
                tag_type = tag.get("type", "")
                print(f"    {tag['code']:12s} {tag['name']:30s} ({tag_type})")

    if show_all or show == "severity":
        print("\n## P 等级（严重度顺序：P0 > P1 > P2）")
        sev = config.get("severity_schema", {})
        for lv in sev.get("levels", []):
            print(f"  {lv}: {sev.get('definitions', {}).get(lv, '')}")

    if show_all or show == "scoring":
        print("\n## 评分规则（0-4 分）")
        scoring = config.get("scoring_rules", {})
        if scoring.get("description"):
            for line in str(scoring["description"]).strip().splitlines():
                print(f"  {line}")

    if show_all or show == "ft_cross":
        print("\n## F×T 交叉校验规则")
        for rule in config.get("ft_cross_validation", {}).get("rules", []):
            scenes = ",".join(rule.get("scenes", []))
            tasks = ",".join(rule.get("tasks", []))
            must = ",".join(rule.get("must_check", []))
            print(f"  - 场景[{scenes}] × 任务[{tasks}]")
            print(f"      必检: {must}")
            print(f"      说明: {rule.get('note', '')}")

    if show_all or show == "families":
        print("\n## 题型族路由（task_family_rules）")
        rules = config.get("task_family_rules", {})
        for fam_name, fam_cfg in rules.items():
            if fam_name.startswith("_"):
                continue
            if family and fam_name != family:
                continue
            print(f"\n  ### {fam_name}  ({fam_cfg.get('display_name', '')})")
            print(f"    匹配 norm_task: {', '.join(fam_cfg.get('match_norm_task', [])) or '(无,通用兜底)'}")
            print(f"    输入特征: {fam_cfg.get('input_features', '')}")
            focus = fam_cfg.get("focus_labels", [])
            if focus:
                print(f"    重点标签: {', '.join(focus)}")
            forbid = fam_cfg.get("forbid_primary", [])
            if forbid:
                print(f"    禁用 primary: {', '.join(forbid)}")
            extra = fam_cfg.get("extra_rules", [])
            if extra:
                print(f"    额外规则:")
                for r in extra:
                    print(f"      - {r}")

    print(f"\n{'='*70}")
    print(f"💡 修改方式: 编辑 config/rvec/{{rules,families,prompts}}.yaml 后重新跑命令验证")
    print(f"   命令示例:")
    print(f"     python fin_rvec_tag.py --show-rules --section scoring")
    print(f"     python fin_rvec_tag.py --show-rules --section families --family QA_CHOICE")
    print(f"{'='*70}\n")


def _cmd_show_prompt(config: Dict, family: str = "QA_CHOICE",
                     sample_record: Optional[Dict] = None) -> None:
    """打印当前生效的 prompt 模板（system + 指定 family 的 user）。

    若提供 sample_record（来自 --with-sample），则额外输出"真实拼接"后的 user prompt，
    让用户能直观看到占位符替换效果。
    """
    schema_text = _build_schema_text(config)
    few_shot_text = _build_few_shot_text(config)
    prompts = _resolve_prompts(config)

    yaml_loaded = bool(config.get("_prompts"))
    print(f"\n{'='*70}")
    print(f"📝 当前生效的 Prompt 模板")
    print(f"   配置来源: {'config/rvec/prompts.yaml（已加载）' if yaml_loaded else '脚本内置常量'}")
    print(f"   Family:   {family}")
    print(f"{'='*70}")

    print(f"\n── 1️⃣  System Prompt 模板（含占位符）──\n")
    print(prompts["system_prompt"])

    print(f"\n── 2️⃣  System Prompt（schema/few-shot 已注入）──\n")
    rendered_system = prompts["system_prompt"].format(
        schema_text="<<schema_text 已省略，可用 --show-rules 查看>>",
        few_shot_text="<<few_shot_text 已省略>>",
    )
    print(rendered_system[:1500] + ("\n...(已截断)" if len(rendered_system) > 1500 else ""))

    print(f"\n── 3️⃣  {family} User Prompt 模板（含占位符）──\n")
    family_tpl = prompts["family_templates"].get(family, "")
    print(family_tpl)

    if sample_record is not None:
        print(f"\n── 4️⃣  {family} User Prompt（用样本数据真实拼接）──\n")
        try:
            rendered_user = _render_user_prompt_for_record(sample_record, family, config, prompts)
            print(rendered_user)
        except Exception as e:
            print(f"⚠️  拼接失败: {e}")

    print(f"\n{'='*70}")
    print(f"💡 修改 prompt: 编辑 config/rvec/prompts.yaml 后重新跑命令验证")
    print(f"   命令示例:")
    print(f"     python fin_rvec_tag.py --show-prompt --family QA_CHOICE")
    print(f"     python fin_rvec_tag.py --show-prompt --family REPORT_EVAL --with-sample data.csv")
    print(f"{'='*70}\n")


def _render_user_prompt_for_record(record: Dict, family: str,
                                    config: Dict, prompts: Dict) -> str:
    """复用 label_one 的字段抽取/family 头/judge 摘要 等逻辑，渲染 user prompt（不调 LLM）。"""
    question = _parse_json_field(_map_field(record, "question"))
    context = _parse_json_field(_map_field(record, "context"))
    answer = _parse_json_field(_map_field(record, "answer"))
    reference = _map_field(record, "reference")
    ground_truth_unstructured = _map_field(record, "ground_truth_unstructured")
    reasoning = _parse_json_field(_map_field(record, "reasoning"))
    accuracy = _map_field(record, "accuracy") or "（无）"
    model_choice = _map_field(record, "model_choice") or "（无）"
    judge_comment = _map_field(record, "judge_comment")
    expected_choice = "（无）"
    try:
        jc = json.loads(judge_comment) if judge_comment else {}
        if isinstance(jc, dict) and jc.get("expected_choice"):
            expected_choice = str(jc["expected_choice"])
    except (json.JSONDecodeError, TypeError):
        pass
    if context and context == question:
        context = ""

    header = _build_family_header(family, config)
    judge_summary = _summarize_judge_comment(judge_comment, family)
    truncation_warning = _build_truncation_warning(record)
    multi_dim_scores = _build_multi_dim_scores(record)

    template = prompts["family_templates"].get(family, prompts["family_templates"].get("GENERIC", ""))
    fmt_kwargs = dict(
        family_display=header["family_display"],
        family_features=header["family_features"],
        family_focus=header["family_focus"],
        family_extra_rules=header["family_extra_rules"],
        question=question[:500] + ("...(已截断仅供预览)" if len(question) > 500 else ""),
        context=(context[:500] + ("...(已截断仅供预览)" if len(context) > 500 else "")) if context else "（无附加题面信息）",
        expected_output=(reference[:500] + ("...(已截断仅供预览)" if len(reference) > 500 else "")) if reference else "（无）",
        ground_truth_unstructured=ground_truth_unstructured or "（无）",
        answer=answer[:500] + ("...(已截断仅供预览)" if len(answer) > 500 else ""),
        reasoning=(reasoning[:300] + ("..." if len(reasoning) > 300 else "")) if reasoning else "（无）",
        accuracy=accuracy,
        model_choice=model_choice,
        expected_choice=expected_choice,
        judge_summary=judge_summary,
        truncation_warning=truncation_warning,
        multi_dim_scores=multi_dim_scores,
    )
    return template.format(**fmt_kwargs)


# ── A/B 对比专用：用指定 config 跑 preview，返回结果列表 ──
_AB_DIFF_FIELDS = [
    "label_fin_scene", "label_task_type",
    "label_rve_primary", "label_rve_all", "label_rve_score_all",
    "label_severity", "label_score",
    "label_highlights",
]


def _run_preview_with_config(records: List[Dict], config_path: Optional[str],
                              llm: LLMClient, n: int) -> List[Dict]:
    """用指定 config 路径加载配置并跑 preview，返回 success 列表。"""
    cfg = _load_config(config_path)
    schema_text = _build_schema_text(cfg)
    few_shot_text = _build_few_shot_text(cfg)
    prompts = _resolve_prompts(cfg)
    sys_prompt = prompts["system_prompt"].format(schema_text=schema_text, few_shot_text=few_shot_text)
    cfg["_family_templates"] = prompts["family_templates"]
    success, _failed = run_batch(records[:n], llm, sys_prompt, cfg, workers=1, retry_limit=1)
    return success


def _build_ab_diff(before_results: List[Dict], after_results: List[Dict],
                   mapping: Dict[str, str]) -> Dict:
    """对齐 before/after 结果，输出聚焦核心字段的 diff。"""
    diffs = []
    n = min(len(before_results), len(after_results))
    for i in range(n):
        b = before_results[i]
        a = after_results[i]
        # 输入摘要：取 question/answer/accuracy
        input_summary = {}
        for role in ("question", "answer", "model_choice", "accuracy"):
            col = mapping.get(role)
            if col:
                v = str(b.get(col, ""))
                input_summary[role] = v[:120] + ("..." if len(v) > 120 else "")
        # 仅 diff 核心标签字段
        before_labels = {k: str(b.get(k, "")) for k in _AB_DIFF_FIELDS}
        after_labels = {k: str(a.get(k, "")) for k in _AB_DIFF_FIELDS}
        diff_fields = [k for k in _AB_DIFF_FIELDS if before_labels.get(k) != after_labels.get(k)]
        diffs.append({
            "_row": i + 1,
            "input_summary": input_summary,
            "before": before_labels,
            "after": after_labels,
            "diff_fields": diff_fields,
        })
    return {
        "mode": "ab_diff",
        "preview_count": n,
        "diff_fields_tracked": _AB_DIFF_FIELDS,
        "comparisons": diffs,
    }


def _cmd_explain_one(record: Dict, llm: LLMClient, system_prompt: str, config: Dict) -> Dict:
    """单条深度调试：展示从原始输入到最终标签的完整 8 段过程。

    输出顺序：
      ① 原始数据快照（用户视角）
      ② 字段抽取（脚本看到了什么）
      ③ Family 路由 + family header
      ④ 完整 user prompt（喂给 LLM 的）
      ⑤ LLM 原始返回 + 解析 JSON
      ⑥ F×T 二次修正（如果触发）
      ⑦ 后处理 diff（_post_process_labels 改了什么）
      ⑧ 最终标签
    """
    # ── 1. 原始快照 ──
    family = classify_family(record, config)
    snap = _build_input_snapshot(record, family)
    print(f"\n{'='*70}")
    print(f"🔬 EXPLAIN 模式：单条深度调试")
    print(f"{'='*70}")
    print(f"\n【1️⃣ 原始数据快照】")
    _print_snapshot(snap)

    # ── 2. 字段抽取 ──
    extracted = {
        "question_len": len(_parse_json_field(_map_field(record, "question"))),
        "context_len": len(_parse_json_field(_map_field(record, "context"))),
        "answer_len": len(_parse_json_field(_map_field(record, "answer"))),
        "reasoning_len": len(_parse_json_field(_map_field(record, "reasoning"))),
        "reference_len": len(_map_field(record, "reference")),
        "judge_comment_len": len(_map_field(record, "judge_comment")),
        "norm_task": _map_field(record, "norm_task"),
        "truncated": _map_field(record, "truncated"),
    }
    print(f"\n【2️⃣ 字段抽取统计】")
    for k, v in extracted.items():
        print(f"  {k}: {v}")

    # ── 3. Family 路由 ──
    rules = config.get("task_family_rules", {}).get(family, {})
    print(f"\n【3️⃣ Family 路由 → {family}】")
    print(f"  display:       {rules.get('display_name', '')}")
    print(f"  focus_labels:  {', '.join(rules.get('focus_labels', [])) or '(无)'}")
    print(f"  forbid_primary:{', '.join(rules.get('forbid_primary', [])) or '(无)'}")
    if rules.get("extra_rules"):
        print(f"  extra_rules:")
        for r in rules["extra_rules"]:
            print(f"    - {r}")

    # ── 4. 完整 user prompt ──
    prompts = _resolve_prompts(config)
    user_msg = _render_user_prompt_for_record(record, family, config, prompts)
    print(f"\n【4️⃣ 喂给 LLM 的完整 USER PROMPT】({len(user_msg)} 字符)")
    print("─" * 70)
    print(user_msg)
    print("─" * 70)

    # ── 5. LLM 调用 ──
    print(f"\n【5️⃣ LLM 调用中...】model={llm.model}")
    try:
        raw, parsed = llm.chat_json_with_raw(system_prompt, user_msg)
    except Exception as e:
        print(f"❌ LLM 调用失败: {e}")
        return {"error": str(e)}
    print(f"  ✅ LLM 原始返回 ({len(raw)} 字符):")
    print("─" * 70)
    print(raw[:3000] + ("\n...(已截断)" if len(raw) > 3000 else ""))
    print("─" * 70)
    print(f"  ✅ 解析后 JSON:")
    print(json.dumps(parsed, ensure_ascii=False, indent=2))

    # ── 6. F×T 二次修正 ──
    accuracy = _map_field(record, "accuracy") or "（无）"
    parsed["_accuracy_hint"] = accuracy
    parsed["_family_hint"] = family
    ft_hint = _get_ft_cross_hint(config, parsed)
    parsed.pop("_accuracy_hint", None)
    parsed.pop("_family_hint", None)
    print(f"\n【6️⃣ F×T 交叉校验】")
    if ft_hint:
        print(f"  触发二次修正：")
        for line in ft_hint.strip().splitlines():
            print(f"    {line}")
        try:
            verify_msg = (
                f"你刚才给出的标注结果如下：\n```json\n{json.dumps(parsed, ensure_ascii=False, indent=2)}\n```\n"
                f"{ft_hint}\n\n请根据交叉校验提醒重新审视标注，输出修正后的完整JSON。"
            )
            raw2, parsed = llm.chat_json_with_raw(system_prompt, verify_msg)
            print(f"  ✅ 二次修正后:")
            print(json.dumps(parsed, ensure_ascii=False, indent=2))
        except Exception as e:
            print(f"  ⚠️ 二次修正失败（保留原结果）: {e}")
    else:
        print(f"  未触发（无交叉校验冲突）")

    # ── 7. 后处理 diff ──
    print(f"\n【7️⃣ 后处理（_post_process_labels）】")
    pre_post = {k: parsed.get(k, "") for k in _AB_DIFF_FIELDS}
    result = dict(record)
    for key in ["label_fin_scene", "label_task_type",
                "label_rve_primary", "label_rve_all", "label_rve_score_all",
                "label_severity", "label_score",
                "label_highlights", "label_evidence", "label_reason"]:
        result[key] = parsed.get(key, "")
    result = _post_process_labels(result, accuracy)
    post_post = {k: result.get(k, "") for k in _AB_DIFF_FIELDS}
    diff_keys = [k for k in _AB_DIFF_FIELDS if str(pre_post.get(k, "")) != str(post_post.get(k, ""))]
    if diff_keys:
        print(f"  后处理改写了 {len(diff_keys)} 个字段：")
        for k in diff_keys:
            print(f"    {k}:")
            print(f"        before: {pre_post.get(k, '')}")
            print(f"        after:  {post_post.get(k, '')}")
    else:
        print(f"  无变化")

    # ── 8. 最终标签 ──
    model_tag = str(getattr(llm, "model", "") or "unknown").strip().replace(" ", "_")
    result["labeler"] = f"fin_rvec_tag@{model_tag}"
    result["review_status"] = "pending"
    print(f"\n【8️⃣ 最终标签】")
    _print_label_result(result)
    print(f"\n{'='*70}\n")

    return result


def _pick_records(records: List[Dict], pick_spec: str) -> List[Dict]:
    """按 --pick 规则筛选记录。

    支持格式：
      "1,3,5"            → 按 1-based 行号
      "id=ABC,DEF"       → 按 ID 字段（answer_id/id/item_id/sample_id）匹配
      "1-5"              → 行号区间
    多种规则可以混用，逗号分隔。
    """
    if not pick_spec:
        return records
    selected = []
    seen_idx = set()
    spec = pick_spec.strip()

    # 拆分 id=xxx 段
    id_values = []
    other_parts = []
    for part in spec.split(";"):
        part = part.strip()
        if part.lower().startswith("id="):
            id_values.extend([v.strip() for v in part[3:].split(",") if v.strip()])
        else:
            other_parts.extend([p.strip() for p in part.split(",") if p.strip()])

    # 按行号
    for token in other_parts:
        if "-" in token:
            try:
                a, b = token.split("-", 1)
                start, end = int(a), int(b)
                for i in range(start, end + 1):
                    idx = i - 1
                    if 0 <= idx < len(records) and idx not in seen_idx:
                        selected.append(records[idx])
                        seen_idx.add(idx)
            except ValueError:
                logger.warning(f"⚠️ --pick 区间解析失败: {token}")
        else:
            try:
                idx = int(token) - 1
                if 0 <= idx < len(records) and idx not in seen_idx:
                    selected.append(records[idx])
                    seen_idx.add(idx)
            except ValueError:
                logger.warning(f"⚠️ --pick 行号解析失败: {token}")

    # 按 ID
    if id_values:
        id_set = set(id_values)
        for i, r in enumerate(records):
            if i in seen_idx:
                continue
            rid = _map_field(r, "id")
            if rid and rid in id_set:
                selected.append(r)
                seen_idx.add(i)

    return selected


# ════════════════════════════════════════════════════════════
# 7. 数据探查
# ════════════════════════════════════════════════════════════

def inspect_data(records: List[Dict], file_path: str):
    if not records:
        print(json.dumps({"error": "数据为空"}, ensure_ascii=False, indent=2))
        return

    columns = list(records[0].keys())
    mapping = _auto_field_mapping(records)
    sample = records[0]

    field_info = {}
    for role, col in mapping.items():
        field_info[role] = {
            "column": col,
            "sample": str(sample.get(col, ""))[:200],
            "non_empty": sum(1 for r in records if str(r.get(col, "")).strip()),
        }

    preview = []
    for i, r in enumerate(records[:3]):
        row = {"_row": i + 1}
        for role, col in mapping.items():
            val = str(r.get(col, ""))
            row[f"{role}({col})"] = val[:120] + ("..." if len(val) > 120 else "")
        preview.append(row)

    label_cols = [c for c in columns if c.startswith("label_")]
    labeled = sum(1 for r in records if any(str(r.get(c, "")).strip() for c in label_cols)) if label_cols else 0

    # ── v3.5 patch ②：字段映射确认（按角色分类，输出固定结构） ──
    field_confirmation = {
        "primary": {  # 核心打标必备
            "题目": mapping.get("question") or mapping.get("context") or "❌ 未识别",
            "模型回答": mapping.get("answer") or "❌ 未识别",
            "参考答案": mapping.get("reference") or mapping.get("ground_truth_unstructured") or "（无）",
            "推理过程": mapping.get("reasoning") or "（无）",
        },
        "auxiliary": {  # 辅助评测线索（不会覆盖最终标签）
            "原始评分": mapping.get("accuracy") or "（无）",
            "judge 评注": mapping.get("judge_comment") or "（无）",
            "题型族字段": mapping.get("norm_task") or "（无 → 默认 GENERIC）",
            "截断标记": mapping.get("truncated") or "（无）",
            "多维评分": [k for k in ("factuality_score", "recall_score", "reasoning_score",
                                     "structure_score", "comprehensive_score") if k in mapping] or "（无）",
        },
        "id": {  # 用于跨次打标对齐
            "ID 列": mapping.get("id") or "（无 → 用行号兜底）",
        },
    }

    # ── v3.5 patch ③：抽样建议 ──
    n = len(records)
    if n <= 20:
        sample_advice = {
            "recommend": "all",
            "reason": f"数据量仅 {n} 条，建议直接全量打标",
            "command_hint": "直接 --auto",
        }
    elif n <= 100:
        sample_advice = {
            "recommend": "all_or_preview",
            "reason": f"数据量 {n} 条（≤100），可直接全量；如需先验证 prompt 效果建议 --preview 3",
            "command_hint": "建议先 --preview 3 看效果，再 --auto 全量",
        }
    else:
        sample_size = min(50, max(20, n // 20))
        sample_advice = {
            "recommend": "sample_first",
            "reason": f"数据量 {n} 条（>100），强烈建议先抽样 {sample_size} 条试跑确认标注口径",
            "command_hint": f"先 --sample-size {sample_size} --auto 抽样跑批，确认标注质量后再 --auto 跑全量",
        }

    # ── v3.5 patch ①：首轮响应所需的就绪状态 ──
    ready_for_labeling = bool(mapping.get("question") and mapping.get("answer"))
    family_dist = {}
    for r in records:
        # _classify_family 在 fin_rvec_tag 内部，这里临时简化：按 norm_task 分组
        nt = _map_field(r, "norm_task") or "GENERIC"
        family_dist[nt] = family_dist.get(nt, 0) + 1
    family_dist = dict(sorted(family_dist.items(), key=lambda x: -x[1])[:5])

    # ── v3.6 加强：字段完整性状态总结（Agent 直接用这个字段决定首轮回复是简短还是详细） ──
    # 主字段缺失 → 阻断打标
    primary_missing = [r for r in ("题目", "模型回答")
                       if field_confirmation["primary"][r] == "❌ 未识别"]
    # 辅助字段缺失 → 不阻断但影响某些维度
    aux_missing = [r for r, v in field_confirmation["auxiliary"].items()
                   if v in ("（无）", "（无 → 默认 GENERIC）")]
    # 推理 / 参考答案缺失 → 影响二级证据
    optional_missing = [r for r in ("参考答案", "推理过程")
                        if field_confirmation["primary"][r] in ("（无）",)]

    if primary_missing:
        completeness_status = "BLOCKED"
        completeness_label = f"❌ 阻断：必需字段缺失 → {', '.join(primary_missing)}"
    elif optional_missing or aux_missing:
        completeness_status = "READY_WITH_GAPS"
        gaps = optional_missing + aux_missing
        completeness_label = f"⚠️  可打标但有缺口（不阻断）：{', '.join(gaps[:3])}"
    else:
        completeness_status = "FULLY_READY"
        completeness_label = "✅ 字段完整，可直接全量打标"

    report = {
        "file": file_path,
        "total_rows": len(records),
        "total_columns": len(columns),
        "columns": columns,
        "ready_for_labeling": ready_for_labeling,
        "completeness_status": completeness_status,
        "completeness_label": completeness_label,
        "missing_required": primary_missing,
        "missing_optional": optional_missing + aux_missing,
        "field_mapping": field_info,
        "field_confirmation": field_confirmation,
        "norm_task_dist": family_dist,
        "sample_advice": sample_advice,
        "recommended_mode": "fin_rvec",
        "data_preview": preview,
        "existing_labels": {"label_columns": label_cols, "labeled": labeled, "unlabeled": len(records) - labeled},
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))


# ════════════════════════════════════════════════════════════
# 8. 主函数
# ════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="金融 RVEC 综合打标（自包含脚本）")
    parser.add_argument("--input", default=None, help="输入文件 (xlsx/csv/jsonl) — 打标/预览/探查必填")
    parser.add_argument("--output", default="output", help="输出目录（默认 output/）")
    parser.add_argument("--config", default=None, help="RVEC 配置文件路径")

    parser.add_argument("--auto", action="store_true",
                        help="自动模式：一条命令完成 探查→打标→输出→manifest（推荐 Agent 使用）")
    parser.add_argument("--inspect", action="store_true", help="数据探查模式（不调用LLM）")
    parser.add_argument("--preview", type=int, nargs="?", const=3, default=None,
                        help="预览模式：打标N条（默认3）并输出对比JSON")
    parser.add_argument("--sample-size", type=int, default=None, help="随机抽样数量")
    parser.add_argument("--filter-family", default=None,
                        choices=["QA_CHOICE", "SENTIMENT", "REPORT_EVAL", "LONG_GEN", "GENERIC"],
                        help="按题型族筛选（在抽样/preview 前生效），便于针对单一 family 验证 prompt 效果")
    parser.add_argument("--pick", default=None,
                        help="挑选特定样本进行预览/explain。格式：'1,3,5' 行号 / '1-5' 区间 / 'id=ABC,DEF' 按ID。可混用：'1,3;id=XYZ'")
    parser.add_argument("--explain", action="store_true",
                        help="深度调试模式：对单条样本输出完整 8 段过程（输入快照→字段抽取→family→user prompt→LLM 原始返回→F×T 修正→后处理 diff→最终标签）")
    parser.add_argument("--snapshot-only", action="store_true",
                        help="仅打印输入快照（不调 LLM），用于核对脚本看到的输入和你以为的是否一致")

    # ── 工作流子命令（不调用 LLM，仅查看规则/prompt）──
    parser.add_argument("--show-rules", action="store_true",
                        help="查看当前生效的规则（RVEC 标签体系/评分规则/family 路由）")
    parser.add_argument("--section", default=None,
                        choices=["all", "rvec", "scoring", "severity", "scenes", "tasks", "ft_cross", "families"],
                        help="--show-rules 时限定要看的章节（默认 all）")
    parser.add_argument("--family", default=None,
                        choices=["QA_CHOICE", "SENTIMENT", "REPORT_EVAL", "LONG_GEN", "GENERIC"],
                        help="--show-rules / --show-prompt 时限定要看的 family")
    parser.add_argument("--show-prompt", action="store_true",
                        help="查看当前生效的 prompt 模板（system + family）")
    parser.add_argument("--with-sample", default=None,
                        help="--show-prompt 时指定一份样本数据，输出真实拼接后的 user prompt")

    # ── A/B 对比 ──
    parser.add_argument("--before", default=None,
                        help="--preview 时指定 before 配置文件，与 --after 配合做改前/改后对比")
    parser.add_argument("--after", default=None,
                        help="--preview 时指定 after 配置文件，与 --before 配合做改前/改后对比")

    parser.add_argument("--endpoint", default=None, choices=["iquest", "zerail"], help="预置端点")
    parser.add_argument("--base-url", default=None)
    parser.add_argument("--api-key", default=None)
    parser.add_argument("--model", default=None)
    parser.add_argument("--temperature", type=float, default=0.0)

    parser.add_argument("--workers", type=int, default=2, help="并发线程数（默认2，避免限流）")
    parser.add_argument("--output-format", default=None, help="输出格式(xlsx/csv/jsonl)，默认跟随输入")
    parser.add_argument("--progress-file", default=None,
                        help="进度文件路径（JSON），打标过程中实时更新，可通过 cat 查看进度")

    args = parser.parse_args()

    # ── 工作流：show-rules / show-prompt（不需要 input/LLM）──
    if args.show_rules:
        config = _load_config(args.config)
        _cmd_show_rules(config, section=args.section or "all", family=args.family)
        return
    if args.show_prompt:
        config = _load_config(args.config)
        sample_record = None
        if args.with_sample:
            recs = read_data(args.with_sample)
            if recs:
                if args.family:
                    matched = [r for r in recs if classify_family(r, config) == args.family]
                    sample_record = matched[0] if matched else recs[0]
                else:
                    sample_record = recs[0]
        _cmd_show_prompt(config, family=args.family or "QA_CHOICE", sample_record=sample_record)
        return

    # 其余模式必须有 --input
    if not args.input:
        parser.error("除 --show-rules / --show-prompt 外，--input 为必填项")

    # ── 读取数据 ──
    logger.info(f"📂 读取数据: {args.input}")
    records = read_data(args.input)
    logger.info(f"📊 共 {len(records)} 行, {len(records[0].keys()) if records else 0} 列")

    # ── 探查模式 ──
    if args.inspect and not args.auto:
        inspect_data(records, args.input)
        return

    # ── auto 模式：先输出探查摘要 ──
    if args.auto:
        mapping = _auto_field_mapping(records)
        label_cols = [c for c in records[0].keys() if c.startswith("label_")] if records else []
        labeled = sum(1 for r in records if any(str(r.get(c, "")).strip() for c in label_cols)) if label_cols else 0
        print(f"📊 数据探查: {len(records)} 行, {len(records[0].keys()) if records else 0} 列")
        print(f"   字段映射: " + ", ".join(f"{k}→{v}" for k, v in mapping.items()))
        print(f"   已标注: {labeled}, 待打标: {len(records) - labeled}")

    # ── 加载配置 ──
    config = _load_config(args.config)

    # ── 按 family 过滤（在抽样之前，便于 preview 单一题型族）──
    if args.filter_family:
        before_n = len(records)
        records = [r for r in records if classify_family(r, config) == args.filter_family]
        logger.info(f"🎯 family 过滤: {args.filter_family} | {before_n} → {len(records)} 条")
        if not records:
            logger.warning(f"⚠️ 没有 family={args.filter_family} 的记录，退出")
            return

    # ── 按 --pick 挑选（行号/区间/ID）──
    if args.pick:
        before_n = len(records)
        records = _pick_records(records, args.pick)
        logger.info(f"🎯 --pick {args.pick} | {before_n} → {len(records)} 条")
        if not records:
            logger.warning(f"⚠️ --pick 未匹配到任何记录，退出")
            return

    # ── 仅快照模式（不调 LLM）──
    if args.snapshot_only:
        n = args.preview if args.preview is not None else min(3, len(records))
        n = min(n, len(records))
        print(f"\n{'='*70}")
        print(f"🔍 输入快照模式（不调 LLM） | 共 {n} 条")
        print(f"{'='*70}")
        for idx, r in enumerate(records[:n], 1):
            family = classify_family(r, config)
            snap = _build_input_snapshot(r, family)
            print(f"\n── 样本 {idx} ──")
            _print_snapshot(snap)
        print(f"\n{'='*70}\n")
        return

    # ── 抽样 ──
    if args.sample_size and args.sample_size < len(records):
        random.seed(42)
        records = random.sample(records, args.sample_size)
        logger.info(f"🎲 随机抽样 {args.sample_size} 条")

    # ── 构建 system prompt ──
    schema_text = _build_schema_text(config)
    few_shot_text = _build_few_shot_text(config)
    prompts = _resolve_prompts(config)
    system_prompt = prompts["system_prompt"].format(schema_text=schema_text, few_shot_text=few_shot_text)
    # 把 family 模板注入到 config，让 label_one 能取到（避开使用全局常量）
    config["_family_templates"] = prompts["family_templates"]

    # ── 连接 LLM ──
    if args.endpoint:
        llm = LLMClient(endpoint_name=args.endpoint, temperature=args.temperature)
    elif args.api_key or args.base_url or os.getenv("OPENAI_API_KEY"):
        llm = LLMClient(api_key=args.api_key, base_url=args.base_url,
                         model=args.model or "gpt-4o", temperature=args.temperature)
    else:
        logger.info("🔍 自动探测 LLM 端点...")
        llm = LLMClient.auto_detect(temperature=args.temperature)

    # ── 输出格式 ──
    if args.output_format:
        out_fmt = args.output_format
    else:
        out_fmt = Path(args.input).suffix.lower().lstrip(".")
        if out_fmt not in ("xlsx", "csv", "jsonl"):
            out_fmt = "jsonl"
        if out_fmt == "xls":
            out_fmt = "xlsx"

    # ── 预览模式 ──
    if args.preview is not None or args.explain:
        # explain 模式默认 n=1（除非用户显式 --preview N）
        if args.explain:
            n = args.preview if args.preview is not None else 1
        else:
            n = args.preview
        n = min(n, len(records))

        # ── EXPLAIN 模式：单条/多条深度调试 ──
        if args.explain:
            for idx, r in enumerate(records[:n], 1):
                if n > 1:
                    print(f"\n\n{'#'*70}")
                    print(f"#  EXPLAIN 第 {idx}/{n} 条")
                    print(f"{'#'*70}")
                _cmd_explain_one(r, llm, system_prompt, config)
            return

        # ── A/B 对比模式：--before/--after 同时给定 ──
        if args.before and args.after:
            logger.info(f"🆎 A/B 对比模式 | 改前: {args.before} | 改后: {args.after} | 样本: {n}")
            before_results = _run_preview_with_config(records, args.before, llm, n)
            after_results = _run_preview_with_config(records, args.after, llm, n)
            mapping = _auto_field_mapping(records)
            ab_report = _build_ab_diff(before_results, after_results, mapping)
            ab_report["before_config"] = args.before
            ab_report["after_config"] = args.after
            ab_report["model"] = llm.model

            # 文本展示
            print(f"\n{'='*70}")
            print(f"🆎 A/B 配置对比 | 模型: {llm.model} | 样本: {n}")
            print(f"   改前: {args.before}")
            print(f"   改后: {args.after}")
            print(f"{'='*70}")
            for c in ab_report["comparisons"]:
                print(f"\n── 样本 {c['_row']} ──")
                q = c["input_summary"].get("question", "")
                print(f"  📝 题目: {q}")
                if c["diff_fields"]:
                    print(f"  🔀 差异字段: {', '.join(c['diff_fields'])}")
                    for f in c["diff_fields"]:
                        print(f"     {f}:")
                        print(f"        改前: {c['before'].get(f, '')}")
                        print(f"        改后: {c['after'].get(f, '')}")
                else:
                    print(f"  ✅ 标签结果完全一致（无差异）")

            print(f"\n{'='*70}")
            print("📊 结构化 JSON（A/B diff）:")
            print(json.dumps(ab_report, ensure_ascii=False, indent=2))
            return

        # ── 普通预览（单 config） ──
        logger.info(f"🔍 预览模式：打标 {n} 条")
        preview_records = records[:n]
        success, failed = run_batch(preview_records, llm, system_prompt, config, workers=1, retry_limit=1)

        # 头部信息
        print(f"\n{'='*70}")
        print(f"📋 预览打标结果 | 模型: {llm.model}")
        print(f"   成功: {len(success)} | 失败: {len(failed)}")
        print(f"   tip: 用 --explain 看单条完整调试过程；用 --before/--after 做 A/B 对比")
        print(f"{'='*70}")

        # 每条样本：上半部分输入快照 / 下半部分打标结果，中间分隔线
        for idx, item in enumerate(success, 1):
            family = classify_family(item, config)
            snap = _build_input_snapshot(item, family)
            print(f"\n┌── 样本 {idx} ──────────────────────────────────────────────────┐")
            print(f"│ � 输入侧（脚本看到的内容）")
            _print_snapshot(snap, indent="│   ")
            print(f"├──────────────────────────────────────────────────────────────")
            print(f"│ 📤 打标结果")
            _print_label_result(item, indent="│   ")
            print(f"└──────────────────────────────────────────────────────────────")

        # 失败样本也要展示输入快照（便于排查）
        if failed:
            print(f"\n⚠️  失败样本（共 {len(failed)} 条）:")
            for idx, item in enumerate(failed, 1):
                family = classify_family(item, config)
                snap = _build_input_snapshot(item, family)
                print(f"\n── 失败 {idx} ──")
                _print_snapshot(snap)

        # 简化的结构化 JSON（仅保留打标关键字段，便于程序解析）
        label_keys = [k for k in (success[0].keys() if success else []) if k.startswith("label_")]
        comparisons = []
        for idx, item in enumerate(success, 1):
            family = classify_family(item, config)
            snap = _build_input_snapshot(item, family)
            # JSON 用纯字符串 key，移除 emoji 前缀便于程序消费
            input_clean = {k.split(" ", 1)[-1] if " " in k else k: v
                           for k, v in snap.items() if k != "__family__"}
            input_clean["family"] = snap.get("__family__", family)
            after = {k: str(item.get(k, "")) for k in label_keys}
            comparisons.append({"_row": idx, "input": input_clean, "labels": after})

        print(f"\n{'='*70}")
        print("📊 结构化 JSON（供程序解析）:")
        print(json.dumps({
            "mode": "fin_rvec", "model": llm.model,
            "preview_count": len(success), "failed_count": len(failed),
            "comparisons": comparisons,
        }, ensure_ascii=False, indent=2))
        return

    # ── 识别已标注记录 ──
    need_label, already_labeled = [], []
    for r in records:
        has = (str(r.get("label_status", "")).strip() == "done"
               or str(r.get("label_fin_scene", "")).strip()
               or str(r.get("label_rve_primary", "")).strip())
        if has:
            r.setdefault("label_status", "done")
            already_labeled.append(r)
        else:
            need_label.append(r)

    if already_labeled:
        logger.info(f"📌 {len(already_labeled)} 条已有标注（保留），待打标: {len(need_label)}")

    if not need_label:
        logger.info("✅ 所有记录已有标注，无需打标")
        return

    # ── 打标 ──
    input_name = Path(args.input).stem
    progress_file = args.progress_file
    if not progress_file:
        # 默认在输出目录生成进度文件
        output_dir_p = Path(args.output)
        output_dir_p.mkdir(parents=True, exist_ok=True)
        progress_file = str(output_dir_p / f"{input_name}_progress.json")

    print(f"\n{'='*60}")
    print(f"🚀 金融 RVEC 打标")
    print(f"📊 待打标: {len(need_label)} | 已标注: {len(already_labeled)} | 并发: {args.workers}")
    print(f"🤖 模型: {llm.model}")
    print(f"📈 进度文件: {progress_file}")
    print(f"   (可在另一终端执行: cat {progress_file}  查看实时进度)")
    print(f"{'='*60}\n")

    success, failed = run_batch(need_label, llm, system_prompt, config,
                                workers=args.workers,
                                retry_limit=config.get("retry_limit", 3),
                                progress_file=progress_file)

    all_results = already_labeled + success
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_file = output_dir / f"{input_name}_rvec_labeled.{out_fmt}"
    fail_file = output_dir / f"{input_name}_rvec_failed.{out_fmt}"

    if all_results:
        write_data(all_results, str(out_file), out_fmt)
    if failed:
        write_data(failed, str(fail_file), out_fmt)

    print(f"\n{'='*60}")
    print(f"🎉 打标完成!")
    print(f"   ✅ 成功: {len(success)}")
    if already_labeled:
        print(f"   📌 保留已标注: {len(already_labeled)}")
    print(f"   ❌ 失败: {len(failed)}")
    if all_results:
        print(f"   📄 输出: {out_file}")
    if failed:
        print(f"   📄 失败: {fail_file}")

    # ── 统计摘要（Agent 展示给用户） ──
    if success:
        scores = [int(r.get("label_score", 0)) for r in success if str(r.get("label_score", "")).isdigit()]
        severity_counts = {}
        scene_counts = {}
        rve_counts = {}
        family_counts = {}
        for r in success:
            sev = str(r.get("label_severity", "NONE"))
            severity_counts[sev] = severity_counts.get(sev, 0) + 1
            # family 分布（重新分类一次以便摘要展示）
            fam = classify_family(r, config)
            family_counts[fam] = family_counts.get(fam, 0) + 1
            for scene in str(r.get("label_fin_scene", "")).split("；"):
                scene = scene.strip()
                if scene:
                    # 取 Fxx 编码
                    code = scene.split(" ")[0] if " " in scene else scene
                    scene_counts[code] = scene_counts.get(code, 0) + 1
            for tag in str(r.get("label_rve_all", "")).split("；"):
                tag = tag.strip()
                if tag and tag != "NONE":
                    # 去掉末尾的 P 等级
                    tag_clean = re.sub(r'：P[012]\s*$', '', tag).strip()
                    code = tag_clean.split(" ")[0] if " " in tag_clean else tag_clean
                    rve_counts[code] = rve_counts.get(code, 0) + 1

        print(f"\n── 📊 打标统计摘要 ──")
        if family_counts:
            print(f"  题型族分布: " + " | ".join(f"{k}:{v}" for k, v in sorted(family_counts.items(), key=lambda x: -x[1])))
        if scores:
            avg_score = sum(scores) / len(scores)
            score_dist = {s: scores.count(s) for s in range(5)}
            print(f"  平均分: {avg_score:.2f}")
            print(f"  分数分布: " + " | ".join(f"{s}分:{c}条" for s, c in score_dist.items() if c > 0))
        if severity_counts:
            print(f"  严重度: " + " | ".join(f"{k}:{v}" for k, v in sorted(severity_counts.items())))
        if rve_counts:
            top_rve = sorted(rve_counts.items(), key=lambda x: -x[1])[:5]
            print(f"  TOP5问题标签: " + " | ".join(f"{k}:{v}" for k, v in top_rve))
        if scene_counts:
            top_scenes = sorted(scene_counts.items(), key=lambda x: -x[1])[:5]
            print(f"  TOP5场景: " + " | ".join(f"{k}:{v}" for k, v in top_scenes))

    print(f"{'='*60}")

    # ── JSON Manifest（Agent 解析用，必须是 stdout 最后一段 JSON） ──
    # v3.5 加强：交付校验 + 样例预览（patch ④⑤）
    def _verify_file(p):
        """返回 (exists, size_bytes)；用于 Agent 交付前校验"""
        try:
            sp = Path(p)
            if sp.exists() and sp.is_file():
                return True, sp.stat().st_size
        except Exception:
            pass
        return False, 0

    out_exists, out_size = _verify_file(out_file) if all_results else (False, 0)
    fail_exists, fail_size = _verify_file(fail_file) if failed else (False, 0)

    # 样例预览：从 success 取前 3 条（patch ⑤ v3.6 加强：让用户看得懂好坏）
    # 新增对照面：参考答案 / 原始评分 / evidence / rve_all / highlights / quality_signals
    sample_preview = []
    for r in success[:3]:
        sample_preview.append({
            "id": _map_field(r, "id") or "",
            "family": classify_family(r, config),
            # 输入对照面（让用户能判断标得对不对）
            "question": _truncate_text(_map_field(r, "question") or _map_field(r, "context"), 150),
            "model_response": _truncate_text(_map_field(r, "answer"), 250),
            "reference": _truncate_text(_map_field(r, "reference") or _map_field(r, "ground_truth_unstructured"), 150),
            "accuracy_origin": _map_field(r, "accuracy") if _map_field(r, "accuracy") not in (None, "") else "",
            # 打标结果
            "label_fin_scene": str(r.get("label_fin_scene", "")),
            "label_task_type": str(r.get("label_task_type", "")),
            "label_rve_primary": str(r.get("label_rve_primary", "")),
            "label_rve_all": str(r.get("label_rve_all", "")),
            "label_severity": str(r.get("label_severity", "")),
            "label_score": r.get("label_score", ""),
            "label_evidence": _truncate_text(r.get("label_evidence", ""), 200),
            "label_highlights": str(r.get("label_highlights", "")),
            "label_reason": _truncate_text(r.get("label_reason", ""), 250),
            # 自动质量信号（Agent 用来判断"这条标得稳不稳"）
            "quality_signals": quality_signals_for_record(r),
        })

    # 整批质量速览（Agent 用来在交付时给 1 行总结）
    quality_summary = {}
    if success:
        align_ok = sum(1 for s in sample_preview if "✅" in str(s["quality_signals"].get("score_align", "")))
        with_ev = sum(1 for s in sample_preview if s["quality_signals"].get("has_evidence"))
        with_rsn = sum(1 for s in sample_preview if s["quality_signals"].get("has_reason"))
        sev_ok = sum(1 for s in sample_preview if s["quality_signals"].get("severity_match") == "✅")
        n_prev = len(sample_preview) or 1
        quality_summary = {
            "preview_n": len(sample_preview),
            "score_alignment_rate": f"{align_ok}/{n_prev}",
            "evidence_coverage": f"{with_ev}/{n_prev}",
            "reason_coverage": f"{with_rsn}/{n_prev}",
            "severity_score_match": f"{sev_ok}/{n_prev}",
        }

    manifest = {
        "status": "completed" if not failed else "partial",
        "input_file": str(args.input),
        "total_rows": len(records),
        "success": len(success),
        "failed": len(failed),
        "already_labeled": len(already_labeled),
        "output_file": str(out_file) if all_results else None,
        "output_file_exists": out_exists,
        "output_file_size_bytes": out_size,
        "output_file_size_human": _human_bytes(out_size),
        "failed_file": str(fail_file) if failed else None,
        "failed_file_exists": fail_exists,
        "failed_file_size_bytes": fail_size,
        "delivery_ready": (out_exists and out_size > 0) if all_results else False,
        "sample_preview": sample_preview,
        "quality_summary": quality_summary,
    }
    if success:
        scores = [int(r.get("label_score", 0)) for r in success if str(r.get("label_score", "")).isdigit()]
        if scores:
            manifest["avg_score"] = round(sum(scores) / len(scores), 2)
            manifest["score_distribution"] = {str(s): scores.count(s) for s in range(5) if scores.count(s) > 0}
    print("\n__MANIFEST_START__")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    print("__MANIFEST_END__")


def _human_bytes(n: int) -> str:
    """人类可读的字节大小"""
    if n < 1024:
        return f"{n}B"
    if n < 1024 * 1024:
        return f"{n/1024:.1f}KB"
    if n < 1024 * 1024 * 1024:
        return f"{n/1024/1024:.2f}MB"
    return f"{n/1024/1024/1024:.2f}GB"


def _truncate_text(s, n: int) -> str:
    """文本截断到 n 字符并加 ..."""
    s = str(s) if s is not None else ""
    return s[:n] + ("..." if len(s) > n else "")


def quality_signals_for_record(rec: dict) -> dict:
    """计算单条打标结果的"标得稳不稳"信号（v3.6 patch ⑤ 增强）。

    返回 dict 包含 4 个信号：
      - score_align: 原始评分 vs LLM label_score 一致性
      - has_evidence: label_evidence 是否非空（有证据更可信）
      - has_reason: label_reason 是否够具体（>=20 字符）
      - severity_match: severity 与 score 是否匹配（P0=>0, P1=>0/1, P2=>1/2, NONE=>3/4）
    """
    sigs = {}
    # 1. 原始评分 vs LLM 评分一致性
    try:
        # _map_field 对 0 会判 falsy，这里直接遍历候选列名取（0 也算合法值）
        acc = None
        for col in ("Accuracy", "accuracy", "score", "评分"):
            if col in rec and rec[col] not in (None, ""):
                acc = rec[col]
                break
        sc = rec.get("label_score", "")
        if acc is not None and str(sc).isdigit():
            acc_f = float(acc)
            sc_i = int(sc)
            if acc_f >= 0.99:
                sigs["score_align"] = "✅ 一致" if sc_i >= 3 else "⚠️ 原始判对但打低分"
            elif acc_f <= 0.01:
                sigs["score_align"] = "✅ 一致" if sc_i <= 2 else "⚠️ 原始判错但打高分"
            else:
                sigs["score_align"] = "中分区无对照"
        else:
            sigs["score_align"] = "无原始评分"
    except Exception:
        sigs["score_align"] = "无法判断"
    # 2. evidence 非空
    sigs["has_evidence"] = bool(str(rec.get("label_evidence", "")).strip())
    # 3. reason 够具体
    rsn = str(rec.get("label_reason", "")).strip()
    sigs["has_reason"] = len(rsn) >= 20
    # 4. severity 与 score 是否匹配
    sev = str(rec.get("label_severity", "")).upper()
    try:
        sc_i = int(rec.get("label_score", -1))
        sev_score_map = {"P0": [0], "P1": [0, 1], "P2": [1, 2], "NONE": [3, 4]}
        expected = sev_score_map.get(sev)
        if expected is None:
            sigs["severity_match"] = "无 severity"
        else:
            sigs["severity_match"] = "✅" if sc_i in expected else f"⚠️ {sev} 但 {sc_i} 分"
    except Exception:
        sigs["severity_match"] = "无法判断"
    return sigs


if __name__ == "__main__":
    main()
